"""
Plating Protocol routes blueprint.
Handles plating protocol creation, storage, and export.
"""
import os
import tempfile
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from state import current_experiment
from state.experiment import update_experiment_plating_protocol

# Create blueprint
plating_protocol_bp = Blueprint('plating_protocol', __name__, url_prefix='/api/experiment')


@plating_protocol_bp.route('/plating-protocol', methods=['GET'])
def get_plating_protocol():
    """Get the saved plating protocol for the current experiment."""
    protocol = current_experiment.get('plating_protocol', None)
    if protocol is None:
        return jsonify({'message': 'No protocol saved', 'protocol': None}), 200
    return jsonify({'message': 'Protocol retrieved', 'protocol': protocol}), 200


@plating_protocol_bp.route('/plating-protocol', methods=['POST'])
def save_plating_protocol():
    """Save plating protocol to experiment state."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Add timestamp
        data['saved_at'] = datetime.now().isoformat()

        # Save to experiment state
        update_experiment_plating_protocol(data)

        return jsonify({'message': 'Protocol saved successfully'}), 200
    except Exception as e:
        return jsonify({'error': f'Failed to save protocol: {str(e)}'}), 500


@plating_protocol_bp.route('/plating-protocol/export', methods=['POST'])
def export_plating_protocol():
    """Export plating protocol to Excel or PDF format."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        protocol = data.get('protocol', {})
        export_format = data.get('format', 'excel')

        if export_format == 'excel':
            return export_to_excel(protocol)
        elif export_format == 'pdf':
            from routes.pdf_export_reportlab import export_to_pdf
            return export_to_pdf(protocol)
        else:
            return jsonify({'error': f'Unsupported format: {export_format}'}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500


def export_to_excel(protocol):
    """Export protocol to Excel format with plate grid visualizations."""
    wb = Workbook()

    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell_fill = PatternFill(start_color='EBF5FF', end_color='EBF5FF', fill_type='solid')

    # Get plate configuration
    plate_type = protocol.get('plate_type', '96')
    plate_config = get_plate_config(plate_type)

    # Sheet 1: Protocol Summary
    ws_summary = wb.create_sheet("Protocol Summary")

    # Context information
    context = protocol.get('context', {})
    ws_summary['A1'] = 'Plating Protocol'
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:H1')

    ws_summary['A3'] = 'ELN:'
    ws_summary['B3'] = context.get('eln', '')
    ws_summary['A4'] = 'Author:'
    ws_summary['B4'] = context.get('author', '')
    ws_summary['A5'] = 'Date:'
    ws_summary['B5'] = context.get('date', datetime.now().strftime('%Y-%m-%d'))
    ws_summary['A6'] = 'Plate Type:'
    ws_summary['B6'] = f'{plate_type}-well'

    # Materials List section
    row = 8
    ws_summary.cell(row=row, column=1, value='Materials List').font = Font(bold=True, size=12)
    row += 1
    mat_list_headers = ['#', 'Alias', 'MW (g/mol)', 'CAS Number', 'Barcode', 'Supplier', 'Cat #']
    for col, header in enumerate(mat_list_headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    experiment_materials = current_experiment.get('materials', [])
    for i, mat in enumerate(experiment_materials):
        row += 1
        mat_row = [
            i + 1,
            mat.get('alias', ''),
            mat.get('molecular_weight', ''),
            mat.get('cas', ''),
            mat.get('barcode', ''),
            mat.get('supplier', ''),
            mat.get('catalog_number', ''),
        ]
        for col, value in enumerate(mat_row, 1):
            cell = ws_summary.cell(row=row, column=col, value=value if value else '')
            cell.border = thin_border

    row += 2

    # Materials summary table (dispensing)
    ws_summary.cell(row=row, column=1, value='Dispensing Summary').font = Font(bold=True, size=12)
    row += 1
    headers = ['#', 'Material', 'Alias', 'MW', 'Method', 'Solvent', 'Vol/Well', 'Excess', 'Conc.', 'Mass (mg)', 'Total Vol.']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    materials = protocol.get('materials', [])
    kit_stock_entries = protocol.get('kit_stock_entries', [])
    kit_role_ids = set(e.get('kit_id', '') for e in kit_stock_entries)

    # Write kit stock entries first to the summary
    item_num = 0
    for entry in kit_stock_entries:
        row += 1
        item_num += 1
        stock = entry.get('stock_solution', {}) or {}
        row_data = [
            item_num,
            entry.get('kit_id', ''),
            entry.get('kit_id', ''),
            '',  # no MW for kit
            'Stock',
            stock.get('solvent_name', '') if stock else '',
            f"{stock.get('amount_per_well_value', '')} {stock.get('amount_per_well_unit', '')}" if stock and stock.get('amount_per_well_value') else '',
            f"{stock.get('excess', '')}%" if stock and stock.get('excess') is not None else '',
            f"{stock.get('concentration_value', '')} {stock.get('concentration_unit', '')}" if stock and stock.get('concentration_value') else '',
            '\u2014',  # mass varies
            f"{stock.get('total_volume_value', '')} {stock.get('total_volume_unit', '')}" if stock and stock.get('total_volume_value') else ''
        ]
        for col, value in enumerate(row_data, 1):
            ws_summary.cell(row=row, column=col, value=str(value) if value else '')

    # Write non-kit materials to the summary
    for idx, material in enumerate(materials, 1):
        # Skip individual kit-stock members (already represented by kit entry)
        if material.get('role_id', '') in kit_role_ids and material.get('dispensing_method') == 'stock':
            continue
        row += 1
        item_num += 1
        stock = material.get('stock_solution', {}) or {}

        # Calculate mass
        mass = None
        total_unit = material.get('total_amount_unit', '\u03bcmol')
        is_solvent = total_unit in ('\u03bcL', 'mL')
        method = material.get('dispensing_method', 'neat')

        if method == 'stock' and stock:
            # Prefer pre-calculated value from buildProtocolData
            pre_calc = material.get('calculated_mass_value')
            if pre_calc is not None:
                try:
                    mass = float(pre_calc)
                except (ValueError, TypeError):
                    pass
            if mass is None:
                mass = calculate_mass(
                    stock.get('concentration_value'),
                    stock.get('concentration_unit', 'M'),
                    stock.get('total_volume_value'),
                    stock.get('total_volume_unit', 'mL'),
                    material.get('molecular_weight')
                )
        elif method == 'neat' and not is_solvent:
            # Neat material: totalAmount (μmol) × MW (g/mol) / 1000 → mg
            pre_calc = material.get('calculated_mass_value')
            if pre_calc is not None:
                try:
                    mass = float(pre_calc)
                except (ValueError, TypeError):
                    pass
            if mass is None:
                mw = material.get('molecular_weight')
                total_amt = material.get('total_amount_value')
                if mw and total_amt:
                    try:
                        mass = float(total_amt) * float(mw) / 1000
                    except (ValueError, TypeError):
                        pass

        if method == 'stock':
            method_label = 'Stock'
        elif is_solvent:
            method_label = 'Solvent'
        else:
            method_label = 'Neat'

        is_cocktail = material.get('is_cocktail', False) or material.get('isCocktail', False)

        row_data = [
            item_num,
            f"{material.get('name', '')} (Premixed)" if is_cocktail else material.get('name', ''),
            material.get('alias', ''),
            '' if is_cocktail else material.get('molecular_weight', ''),
            method_label,
            stock.get('solvent_name', '') if stock else '',
            f"{stock.get('amount_per_well_value', '')} {stock.get('amount_per_well_unit', '')}" if stock and stock.get('amount_per_well_value') else '',
            f"{stock.get('excess', '')}%" if stock and stock.get('excess') is not None else '',
            '' if is_cocktail else (f"{stock.get('concentration_value', '')} {stock.get('concentration_unit', '')}" if stock and stock.get('concentration_value') else ''),
            '' if is_cocktail else (f"{mass:.2f}" if mass else ''),
            f"{stock.get('total_volume_value', '')} {stock.get('total_volume_unit', '')}" if stock and stock.get('total_volume_value') else ''
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in [1, 5, 6] else 'left')

        if is_cocktail:
            # Pre-calculate cocktail's amountPerWell in μL for component concentrations
            cocktail_vpw_val = stock.get('amount_per_well_value') if stock else None
            cocktail_vpw_unit = stock.get('amount_per_well_unit', '\u03bcL') if stock else '\u03bcL'
            try:
                cocktail_vpw_ul = float(cocktail_vpw_val) * 1000 if cocktail_vpw_unit == 'mL' else float(cocktail_vpw_val) if cocktail_vpw_val else None
            except (ValueError, TypeError):
                cocktail_vpw_ul = None
            cocktail_excess = float(stock.get('excess', 0)) if stock and stock.get('excess') is not None else 0

            for comp in material.get('components', []):
                row += 1
                c_stock = comp.get('stockSolution') or {}
                c_solvent = c_stock.get('solvent') or {}

                # Concentration: prefer pre-calculated value from buildProtocolData
                c_conc_str = ''
                c_conc_val = comp.get('calculated_concentration_value')
                if c_conc_val is not None:
                    try:
                        c_conc_val = float(c_conc_val)
                        c_conc_str = f"{c_conc_val * 1000:.2f} mM" if c_conc_val < 0.1 else f"{c_conc_val:.3f} M"
                    except (ValueError, TypeError):
                        pass
                if not c_conc_str and cocktail_vpw_ul:
                    # Calculate from component's wellAmounts + cocktail's amountPerWell
                    c_well_amounts = comp.get('well_amounts') or comp.get('wellAmounts') or {}
                    if c_well_amounts:
                        try:
                            vals = [float(v.get('value', 0)) for v in c_well_amounts.values() if v.get('value')]
                            if vals:
                                min_val = min(vals)
                                c_conc_m = min_val / cocktail_vpw_ul
                                c_conc_str = f"{c_conc_m * 1000:.2f} mM" if c_conc_m < 0.1 else f"{c_conc_m:.3f} M"
                        except (ValueError, TypeError):
                            pass
                if not c_conc_str:
                    # Final fallback: stored concentration
                    c_conc = c_stock.get('concentration') or {}
                    c_conc_str = f"{c_conc.get('value', '')} {c_conc.get('unit', 'M')}" if c_conc.get('value') else ''

                # Mass: prefer top-level calculated_mass_value, then legacy fields, then compute
                c_mass_val = comp.get('calculated_mass_value')
                if c_mass_val is None:
                    c_mass_val = (comp.get('calculatedMass') or {}).get('value')
                if c_mass_val is None:
                    c_mass_val = (c_stock.get('calculatedMass') or {}).get('value')
                if c_mass_val is None:
                    comp_mw = comp.get('molecular_weight')
                    comp_total = comp.get('total_amount_value') or (comp.get('totalAmount') or {}).get('value')
                    if comp_mw and comp_total:
                        try:
                            c_mass_val = float(comp_total) * float(comp_mw) * (1 + cocktail_excess / 100) / 1000
                        except (ValueError, TypeError):
                            pass
                try:
                    c_mass_str = f"{float(c_mass_val):.2f}" if c_mass_val is not None else ''
                except (ValueError, TypeError):
                    c_mass_str = ''

                comp_row_data = [
                    '',
                    f"  ↳ {comp.get('name', '')}",
                    comp.get('alias', ''),
                    comp.get('molecular_weight', ''),
                    '',
                    c_solvent.get('name', ''),
                    '',
                    '',
                    c_conc_str,
                    c_mass_str,
                    ''
                ]
                for col, value in enumerate(comp_row_data, 1):
                    cell = ws_summary.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center' if col in [1, 5, 6] else 'left')

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 5
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 10
    ws_summary.column_dimensions['F'].width = 15
    ws_summary.column_dimensions['G'].width = 12
    ws_summary.column_dimensions['H'].width = 10
    ws_summary.column_dimensions['I'].width = 12
    ws_summary.column_dimensions['J'].width = 12
    ws_summary.column_dimensions['K'].width = 12

    # ── Protocol Steps (2 columns) ──
    row += 3
    ws_summary.cell(row=row, column=1, value='Protocol Steps')
    ws_summary.cell(row=row, column=1).font = title_font
    row += 1

    operations = protocol.get('operations', [])
    col1_ops = [(i, op) for i, op in enumerate(operations) if i % 2 == 0]
    col2_ops = [(i, op) for i, op in enumerate(operations) if i % 2 != 0]

    # Helper to style step cells
    def write_step(r, c, idx, op):
        step_num = idx + 1
        op_type = op.get('type', '')
        
        # Format the label based on operation type
        if op_type == 'dispense':
            mat_idx = op.get('materialIndex')
            if mat_idx is not None and mat_idx < len(materials):
                mat = materials[mat_idx]
                name = mat.get('alias') or mat.get('name', 'Material')
                label = f"Dispense: {name}"
            else:
                label = "Dispense"
        elif op_type == 'kit':
            label = op.get('label') or op.get('kitId', 'Kit')
        elif op_type in ['stir', 'wait']:
            label = op.get('label') or op_type.capitalize()
            duration = op.get('duration')
            if duration:
                label = f"{label} ({duration}m)"
        else:
            label = op.get('label') or op_type.capitalize()

        # Step number badge
        ws_summary.cell(row=r, column=c, value=f"{step_num}")
        ws_summary.cell(row=r, column=c).font = Font(bold=True)
        ws_summary.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
        
        if op_type == 'dispense':
            ws_summary.cell(row=r, column=c).fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            ws_summary.cell(row=r, column=c).font = Font(bold=True, color='FFFFFF')
        elif op_type == 'kit':
            ws_summary.cell(row=r, column=c).fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
            ws_summary.cell(row=r, column=c).font = Font(bold=True, color='FFFFFF')
        else:
            ws_summary.cell(row=r, column=c).fill = PatternFill(start_color='9CA3AF', end_color='9CA3AF', fill_type='solid')
            ws_summary.cell(row=r, column=c).font = Font(bold=True, color='FFFFFF')

        # Step Description
        ws_summary.cell(row=r, column=c+1, value=label)
        ws_summary.cell(row=r, column=c+1).alignment = Alignment(vertical='center')

    for i in range(len(col1_ops)):
        # Column 1 (A-B)
        write_step(row, 1, col1_ops[i][0], col1_ops[i][1])
        # Column 2 (D-E)
        if i < len(col2_ops):
            write_step(row, 4, col2_ops[i][0], col2_ops[i][1])
        row += 1

    # ── Helper: write a plate grid to a worksheet ──
    def _write_plate_grid(ws, grid_start_row, well_amounts, plate_config,
                          header_font, thin_border, header_fill, cell_fill):
        ws.cell(row=grid_start_row - 1, column=1, value='Plate positions:')
        ws.cell(row=grid_start_row - 1, column=1).font = Font(bold=True)

        # Column headers
        for col_idx, col_num in enumerate(plate_config['columns'], 2):
            cell = ws.cell(row=grid_start_row, column=col_idx, value=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.fill = header_fill

        # Row headers and values
        for row_idx, row_letter in enumerate(plate_config['rows'], 1):
            cell = ws.cell(row=grid_start_row + row_idx, column=1, value=row_letter)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.fill = header_fill

            for col_idx, col_num in enumerate(plate_config['columns'], 2):
                well_id = f"{row_letter}{col_num}"
                well_data = well_amounts.get(well_id, {})
                value = well_data.get('value', '') if isinstance(well_data, dict) else ''

                cell = ws.cell(row=grid_start_row + row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

                if value:
                    try:
                        num_value = float(value)
                        if num_value >= 100:
                            cell.value = int(num_value)
                        elif num_value >= 10:
                            cell.value = round(num_value, 0)
                        elif num_value >= 1:
                            cell.value = round(num_value, 1)
                        else:
                            cell.value = round(num_value, 2)
                    except (ValueError, TypeError):
                        cell.value = value
                    cell.fill = cell_fill
                else:
                    cell.value = '-'

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, len(plate_config['columns']) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8

    # ── Build per-material / per-kit sheets ──
    processed_kit_ids = set()
    operations = protocol.get('operations', [])

    for idx, op in enumerate(operations):
        op_type = op.get('type')
        step_num = idx + 1

        if op_type == 'kit':
            role_id = op.get('kitId', '')
            if not role_id or role_id in processed_kit_ids:
                continue
            processed_kit_ids.add(role_id)

            safe_name = role_id[:20].translate(str.maketrans(r'\/*?:[]', '_______'))
            sheet_name = f"Step {step_num} - {safe_name}"
            ws_kit = wb.create_sheet(sheet_name)

            # Kit header
            ws_kit['A1'] = f"Step {step_num}: {role_id}"
            ws_kit['A1'].font = title_font
            ws_kit.merge_cells('A1:L1')
            
            kit_members = [m for m in materials if m.get('role_id', '') == role_id]
            if not kit_members:
                continue

            # Dispensing method
            first_mat_method = kit_members[0].get('dispensing_method', 'neat')
            ws_kit['A3'] = 'Method:'
            ws_kit['B3'] = 'Stock Solution' if first_mat_method == 'stock' else 'Neat'

            # List kit members
            member_names = [m.get('alias') or m.get('name', '') for m in kit_members]
            ws_kit['A4'] = 'Members:'
            ws_kit['B4'] = ', '.join(member_names)

            info_row = 5
            stock = kit_members[0].get('stock_solution', {}) or {}
            if first_mat_method == 'stock' and stock:
                ws_kit.cell(row=info_row, column=1, value='Solvent:')
                ws_kit.cell(row=info_row, column=2, value=stock.get('solvent_name', ''))
                info_row += 1

                ws_kit.cell(row=info_row, column=1, value='Volume per well:')
                ws_kit.cell(row=info_row, column=2, value=f"{stock.get('amount_per_well_value', '')} {stock.get('amount_per_well_unit', '')}")
                info_row += 1

                ws_kit.cell(row=info_row, column=1, value='Excess:')
                ws_kit.cell(row=info_row, column=2, value=f"{stock.get('excess', '')}%")
                info_row += 1

                # Find kit-level stock entry for aggregated info
                kit_entry = next((e for e in kit_stock_entries if e.get('kit_id') == role_id), None)
                if kit_entry:
                    ks = kit_entry.get('stock_solution', {}) or {}
                    conc_val = ks.get('concentration_value')
                    if conc_val is not None:
                        ws_kit.cell(row=info_row, column=1, value='Concentration:')
                        ws_kit.cell(row=info_row, column=2, value=f"{conc_val} {ks.get('concentration_unit', 'M')}")
                        info_row += 1
                    tv = ks.get('total_volume_value')
                    if tv is not None:
                        ws_kit.cell(row=info_row, column=1, value='Total volume:')
                        ws_kit.cell(row=info_row, column=2, value=f"{tv} {ks.get('total_volume_unit', 'mL')}")
                        info_row += 1

            # Merge well_amounts from all kit members to show combined positions
            merged_wells = {}
            for km in kit_members:
                for well_id, well_data in (km.get('well_amounts') or {}).items():
                    if well_id not in merged_wells and isinstance(well_data, dict) and well_data.get('value'):
                        merged_wells[well_id] = well_data

            grid_start_row = info_row + 2
            _write_plate_grid(ws_kit, grid_start_row, merged_wells, plate_config,
                              header_font, thin_border, header_fill, cell_fill)

            # Unit note
            unit_row = grid_start_row + len(plate_config['rows']) + 2
            unit = stock.get('amount_per_well_unit', 'μL') if first_mat_method == 'stock' else kit_members[0].get('total_amount_unit', 'μmol')
            ws_kit.cell(row=unit_row, column=1, value=f'Amounts in {unit}')

        elif op_type == 'dispense':
            mat_idx = op.get('materialIndex')
            if mat_idx is None or mat_idx >= len(materials):
                continue
            material = materials[mat_idx]
            
            # Since kit components individual dispense falls under kit, verify if it was individually queued or ignore
            # In our data structure, kit ops wrap their members. Individual dispense is its own op.
            
            # Non-kit material — individual sheet
            raw_name = (material.get('alias') or material.get('name', 'Material'))[:20]
            safe_name = raw_name.translate(str.maketrans(r'\/*?:[]', '_______'))
            sheet_name = f"Step {step_num} - {safe_name}"
            ws_material = wb.create_sheet(sheet_name)

            # Material header
            ws_material['A1'] = f"Step {step_num}: {material.get('alias') or material.get('name', '')}"
            ws_material['A1'].font = title_font
            ws_material.merge_cells('A1:L1')

            # Material info
            mat_total_unit = material.get('total_amount_unit', 'μmol')
            mat_is_solvent = mat_total_unit in ('μL', 'mL')
            ws_material['A3'] = 'Method:'
            if material.get('dispensing_method') == 'stock':
                ws_material['B3'] = 'Stock Solution'
            elif mat_is_solvent:
                ws_material['B3'] = 'Solvent (Neat)'
            else:
                ws_material['B3'] = 'Neat'

            stock = material.get('stock_solution', {}) or {}
            if material.get('dispensing_method') == 'stock' and stock:
                ws_material['A4'] = 'Solvent:'
                ws_material['B4'] = stock.get('solvent_name', '')

                ws_material['A5'] = 'Volume per well:'
                ws_material['B5'] = f"{stock.get('amount_per_well_value', '')} {stock.get('amount_per_well_unit', '')}"

                ws_material['A6'] = 'Excess:'
                ws_material['B6'] = f"{stock.get('excess', '')}%"

                ws_material['A7'] = 'Concentration:'
                ws_material['B7'] = f"{stock.get('concentration_value', '')} {stock.get('concentration_unit', '')}"

                mass = calculate_mass(
                    stock.get('concentration_value'),
                    stock.get('concentration_unit', 'M'),
                    stock.get('total_volume_value'),
                    stock.get('total_volume_unit', 'mL'),
                    material.get('molecular_weight')
                )
                ws_material['A8'] = 'Mass to weigh:'
                ws_material['B8'] = f"{mass:.2f} mg" if mass else '--'

                ws_material['A9'] = 'Total volume:'
                ws_material['B9'] = f"{stock.get('total_volume_value', '')} {stock.get('total_volume_unit', '')}"

            # Plate grid visualization
            grid_start_row = 12
            well_amounts = material.get('well_amounts', {})
            _write_plate_grid(ws_material, grid_start_row, well_amounts, plate_config,
                              header_font, thin_border, header_fill, cell_fill)

            # Add unit note
            unit_row = grid_start_row + len(plate_config['rows']) + 2
            if material.get('dispensing_method') == 'stock' and stock:
                unit = stock.get('amount_per_well_unit', 'μL')
            else:
                unit = material.get('total_amount_unit', 'μmol')
            ws_material.cell(row=unit_row, column=1, value=f'Amounts in {unit}')

    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    # Generate filename
    eln = protocol.get('context', {}).get('eln', 'Protocol')
    date = datetime.now().strftime('%Y-%m-%d')
    filename = f'Plating_Protocol_{eln}_{date}.xlsx'

    return send_file(tmp_path, as_attachment=True, download_name=filename)


def get_plate_config(plate_type):
    """Get plate configuration for given plate type."""
    if plate_type == '24':
        return {
            'rows': ['A', 'B', 'C', 'D'],
            'columns': list(range(1, 7))
        }
    elif plate_type == '48':
        return {
            'rows': ['A', 'B', 'C', 'D', 'E', 'F'],
            'columns': list(range(1, 9))
        }
    else:  # 96
        return {
            'rows': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
            'columns': list(range(1, 13))
        }


def calculate_mass(conc_value, conc_unit, vol_value, vol_unit, molecular_weight):
    """Calculate mass in mg from concentration, volume, and molecular weight."""
    if not conc_value or not vol_value or not molecular_weight:
        return None

    try:
        conc_value = float(conc_value)
        vol_value = float(vol_value)
        molecular_weight = float(molecular_weight)

        # Convert concentration to mol/L
        if conc_unit == 'mM':
            conc_mol_per_l = conc_value / 1000
        elif conc_unit == 'mg/mL':
            # mg/mL is numerically equal to g/L; dividing by MW (g/mol) gives mol/L directly
            conc_mol_per_l = conc_value / molecular_weight
        else:  # M
            conc_mol_per_l = conc_value

        # Convert volume to L
        if vol_unit == 'mL':
            vol_l = vol_value / 1000
        else:  # L
            vol_l = vol_value

        # Mass in mg = C (mol/L) × V (L) × MW (g/mol) × 1000 (mg/g)
        mass_mg = conc_mol_per_l * vol_l * molecular_weight * 1000
        return mass_mg

    except (ValueError, TypeError):
        return None
