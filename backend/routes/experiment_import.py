"""
Import routes blueprint.
Handles experiment data import from Excel format.
"""
import os
import tempfile
from datetime import datetime
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from state import current_experiment
from validation import (
    validate_request, validate_response,
    ExperimentContextSchema, MaterialsListSchema, ProcedureListSchema,
    ProcedureSettingsSchema, AnalyticalDataSchema, ResultsSchema
)

# Create blueprint
import_bp = Blueprint('import', __name__, url_prefix='/api/experiment')

@import_bp.route('/import', methods=['POST'])
def import_experiment():
    """Import experiment data from Excel format"""
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
        
        # Check file size (limit to 10MB)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({'error': 'File size exceeds 10MB limit'}), 400
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Load the workbook
            wb = load_workbook(tmp_path, data_only=True)
            
            # Initialize import results
            import_results = {
                'context': {'imported': False, 'data': {}},
                'materials': {'imported': False, 'count': 0, 'data': []},
                'procedure': {'imported': False, 'count': 0, 'data': []},
                'procedure_settings': {'imported': False, 'data': {}},
                'analytical_data': {'imported': False, 'count': 0, 'data': []},
                'results': {'imported': False, 'count': 0, 'data': []},
                'errors': []
            }
            
            # Import Context sheet
            context_sheet_names = [name for name in wb.sheetnames if 'Context' in name]
            if context_sheet_names:
                try:
                    context_data = import_context_sheet(wb[context_sheet_names[0]])
                    if context_data:
                        # Validate context data
                        try:
                            from validation.utils import validate_data
                            from validation.schemas import ExperimentContextSchema
                            schema = ExperimentContextSchema()
                            validated_context, errors = validate_data(
                                schema, context_data, strict_mode=False,
                                endpoint="Import Context"
                            )
                            if errors:
                                import_results['errors'].extend([f"Context validation: {err}" for err in errors])
                            # Keep original context_data to preserve string date format
                            # Only use validated data for validation warnings
                        except Exception as validation_error:
                            import_results['errors'].append(f"Context validation error: {str(validation_error)}")
                        
                        import_results['context']['imported'] = True
                        import_results['context']['data'] = context_data
                        current_experiment['context'] = context_data
                except Exception as e:
                    import_results['errors'].append(f"Context import error: {str(e)}")
            
            # Import Materials sheet
            if 'Materials' in wb.sheetnames:
                try:
                    materials_data = import_materials_sheet(wb['Materials'])
                    if materials_data:
                        # Validate materials data
                        try:
                            from validation.utils import validate_data
                            from validation.schemas import MaterialSchema
                            schema = MaterialSchema()
                            validated_materials = []
                            for i, material in enumerate(materials_data):
                                validated_material, errors = validate_data(
                                    schema, material, strict_mode=False,
                                    endpoint=f"Import Material {i+1}"
                                )
                                if errors:
                                    import_results['errors'].extend([f"Material {i+1} validation: {err}" for err in errors])
                                validated_materials.append(validated_material)
                            materials_data = validated_materials
                        except Exception as validation_error:
                            import_results['errors'].append(f"Materials validation error: {str(validation_error)}")
                        
                        import_results['materials']['imported'] = True
                        import_results['materials']['count'] = len(materials_data)
                        import_results['materials']['data'] = materials_data
                        current_experiment['materials'] = materials_data
                except Exception as e:
                    import_results['errors'].append(f"Materials import error: {str(e)}")
            
            # Import Design sheet (procedure data)
            design_sheet_names = [name for name in wb.sheetnames if 'Design' in name]
            if design_sheet_names:
                try:
                    procedure_data = import_procedure_sheet(wb[design_sheet_names[0]])
                    if procedure_data:
                        import_results['procedure']['imported'] = True
                        import_results['procedure']['count'] = len(procedure_data)
                        import_results['procedure']['data'] = procedure_data
                        current_experiment['procedure'] = procedure_data
                except Exception as e:
                    import_results['errors'].append(f"Procedure import error: {str(e)}")
            
            # Import Procedure Settings sheet
            procedure_settings_sheets = [name for name in wb.sheetnames if 'Procedure' in name]
            if procedure_settings_sheets:
                try:
                    settings_data = import_procedure_settings_sheet(wb[procedure_settings_sheets[0]])
                    if settings_data:
                        import_results['procedure_settings']['imported'] = True
                        import_results['procedure_settings']['data'] = settings_data
                        current_experiment['procedure_settings'] = settings_data
                except Exception as e:
                    import_results['errors'].append(f"Procedure Settings import error: {str(e)}")
            
            # Import Analytical data sheet
            analytical_sheet_names = [name for name in wb.sheetnames if 'Analytical data' in name or 'Analytical Data' in name]
            if analytical_sheet_names:
                try:
                    analytical_data = import_analytical_sheet(wb[analytical_sheet_names[0]])
                    if analytical_data:
                        import_results['analytical_data']['imported'] = True
                        import_results['analytical_data']['count'] = len(analytical_data.get('data', []))
                        import_results['analytical_data']['data'] = analytical_data
                        
                        # Store analytical data in the expected format (same as upload)
                        if 'analytical_data' not in current_experiment:
                            current_experiment['analytical_data'] = {}
                        
                        # If analytical_data is a list (old format), convert it to new format
                        if isinstance(current_experiment['analytical_data'], list):
                            old_uploads = current_experiment['analytical_data']
                            current_experiment['analytical_data'] = {
                                'selectedCompounds': [],
                                'uploadedFiles': old_uploads
                            }
                        
                        if 'uploadedFiles' not in current_experiment['analytical_data']:
                            current_experiment['analytical_data']['uploadedFiles'] = []
                        
                        current_experiment['analytical_data']['uploadedFiles'].append(analytical_data)
                except Exception as e:
                    import_results['errors'].append(f"Analytical data import error: {str(e)}")
            
            # Import Results sheet
            results_sheet_names = [name for name in wb.sheetnames if 'Results' in name]
            if results_sheet_names:
                try:
                    results_data = import_results_sheet(wb[results_sheet_names[0]])
                    if results_data:
                        import_results['results']['imported'] = True
                        import_results['results']['count'] = len(results_data)
                        import_results['results']['data'] = results_data
                        current_experiment['results'] = results_data
                except Exception as e:
                    import_results['errors'].append(f"Results import error: {str(e)}")
            
            # Check if any data was imported
            any_imported = any([
                import_results['context']['imported'],
                import_results['materials']['imported'],
                import_results['procedure']['imported'],
                import_results['procedure_settings']['imported'],
                import_results['analytical_data']['imported'],
                import_results['results']['imported']
            ])
            
            if not any_imported:
                return jsonify({
                    'error': 'No valid experiment data found in the Excel file. Please ensure the file contains the expected sheets (Context, Materials, Procedure, etc.)'
                }), 400
            
            # Generate summary message
            summary_parts = []
            if import_results['context']['imported']:
                summary_parts.append("Context")
            if import_results['materials']['imported']:
                summary_parts.append(f"{import_results['materials']['count']} materials")
            if import_results['procedure']['imported']:
                summary_parts.append(f"{import_results['procedure']['count']} procedure entries")
            if import_results['procedure_settings']['imported']:
                summary_parts.append("Procedure settings")
            if import_results['analytical_data']['imported']:
                summary_parts.append(f"{import_results['analytical_data']['count']} analytical entries")
            if import_results['results']['imported']:
                summary_parts.append(f"{import_results['results']['count']} results")
            
            summary_message = f"Successfully imported: {', '.join(summary_parts)}"
            if import_results['errors']:
                summary_message += f". Warnings: {'; '.join(import_results['errors'])}"
            
            return jsonify({
                'message': summary_message,
                'import_results': import_results
            })
            
        finally:
            # Clean up temporary file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
                
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'}), 500

def import_context_sheet(ws):
    """Import context data from Context sheet"""
    context_data = {}
    
    # Read context data from rows
    for row in ws.iter_rows(values_only=True):
        if row[0] and len(row) > 1 and row[1] is not None:
            key = str(row[0]).lower().strip()
            
            if key == 'author':
                context_data['author'] = str(row[1]).strip()
            elif key == 'date':
                # Always set today's date when importing, in YYYY-MM-DD
                from datetime import datetime
                context_data['date'] = datetime.now().strftime('%Y-%m-%d')
            elif key == 'project':
                context_data['project'] = str(row[1]).strip()
            elif key in ['eln', 'eln number']:
                context_data['eln'] = str(row[1]).strip()
            elif key == 'objective':
                context_data['objective'] = str(row[1]).strip()
    
    return context_data

def import_materials_sheet(ws):
    """Import materials data from Materials sheet"""
    materials = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip())
    
    # Read materials data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        material = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'nr':
                    continue  # Skip row number
                elif header == 'chemical_name':
                    material['name'] = str(value).strip()
                elif header == 'alias':
                    material['alias'] = str(value).strip()
                elif header == 'cas_number':
                    material['cas'] = str(value).strip()
                elif header == 'molecular_weight':
                    material['molecular_weight'] = str(value).strip()
                elif header == 'smiles':
                    material['smiles'] = str(value).strip()
                elif header == 'barcode':
                    material['barcode'] = str(value).strip()
                elif header == 'role':
                    material['role'] = str(value).strip()
                elif header == 'role_id':
                    material['role_id'] = str(value).strip()
                elif header == 'source':
                    material['source'] = str(value).strip()
                elif header == 'supplier':
                    material['supplier'] = str(value).strip()
        
        # Only add material if it has a name
        if material.get('name'):
            materials.append(material)
    
    return materials

def import_procedure_sheet(ws):
    """Import procedure data from Design sheet (long format).

    Expected columns: well, material_nr, material_alias, amount, dispense_order
    Each row represents one material in one well.
    """
    from collections import defaultdict

    procedure = []
    headers = []

    # Read headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip().lower())

    # Validate expected headers
    if 'well' not in headers:
        return procedure  # Not a valid Design sheet

    # Build column index map
    col_idx = {h: i for i, h in enumerate(headers)}
    well_col = col_idx.get('well')
    nr_col = col_idx.get('material_nr')
    alias_col = col_idx.get('material_alias')
    amount_col = col_idx.get('amount')
    order_col = col_idx.get('dispense_order')

    # Build materials lookup from already-imported Materials list
    # material_nr is 1-based index in the Materials sheet
    materials_list = current_experiment.get('materials', [])
    nr_to_material = {}
    for i, mat in enumerate(materials_list):
        nr_to_material[i + 1] = mat  # 1-based

    # Also build name/alias lookup for fallback matching
    material_by_name = {}
    material_by_alias = {}
    for mat in materials_list:
        name = (mat.get('name', '') or '').strip().lower()
        alias = (mat.get('alias', '') or '').strip().lower()
        if name:
            material_by_name[name] = mat
        if alias:
            material_by_alias[alias] = mat

    # Read rows and group by well
    wells_data = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        def _cell(idx):
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ''

        well = _cell(well_col)
        if not well:
            continue

        # Parse material_nr
        mat_nr_str = _cell(nr_col)
        try:
            mat_nr = int(float(mat_nr_str)) if mat_nr_str else None
        except (ValueError, TypeError):
            mat_nr = None

        alias_val = _cell(alias_col)
        amount_str = _cell(amount_col)
        order_str = _cell(order_col)

        # Parse amount
        try:
            amount = float(amount_str) if amount_str else 0.0
        except (ValueError, TypeError):
            amount = 0.0

        # Parse dispense_order
        try:
            order = int(float(order_str)) if order_str else 999
        except (ValueError, TypeError):
            order = 999

        # Resolve full material properties
        matched_material = None
        if mat_nr and mat_nr in nr_to_material:
            matched_material = nr_to_material[mat_nr]
        elif alias_val:
            # Fallback: try matching by alias/name
            alias_lower = alias_val.lower()
            matched_material = (
                material_by_alias.get(alias_lower)
                or material_by_name.get(alias_lower)
            )

        if matched_material:
            role = (matched_material.get('role', '') or '').lower()
            unit = 'μL' if role == 'solvent' else 'μmol'

            entry = {
                'name': matched_material.get('name', alias_val),
                'alias': matched_material.get('alias', alias_val),
                'cas': matched_material.get('cas', ''),
                'molecular_weight': matched_material.get('molecular_weight', ''),
                'barcode': matched_material.get('barcode', ''),
                'smiles': matched_material.get('smiles', ''),
                'role': matched_material.get('role', ''),
                'source': matched_material.get('source', ''),
                'amount': amount,
                'unit': unit,
                '_order': order,
            }
        else:
            # Material not found in Materials list — use what we have
            entry = {
                'name': alias_val,
                'alias': alias_val,
                'amount': amount,
                'unit': 'μmol',
                '_order': order,
            }

        wells_data[well].append(entry)

    # Convert grouped data to procedure list, sorted by well position
    def _well_sort_key(well):
        row_letter = well[0].upper()
        try:
            col_num = int(well[1:])
        except ValueError:
            col_num = 99
        return (row_letter, col_num)

    for well in sorted(wells_data.keys(), key=_well_sort_key):
        entries = wells_data[well]
        # Sort by dispense_order within each well
        entries.sort(key=lambda e: e.get('_order', 999))
        # Remove internal _order key
        for e in entries:
            e.pop('_order', None)
        procedure.append({
            'well': well,
            'materials': entries,
        })

    return procedure


def import_procedure_settings_sheet(ws):
    """Import procedure settings from Procedure Settings sheet"""
    settings = {
        'reactionConditions': {
            'temperature': '',
            'time': '',
            'pressure': '',
            'wavelength': '',
            'remarks': ''
        },
        'analyticalDetails': {
            'uplcNumber': '',
            'method': '',
            'duration': '',
            'wavelength': '',
            'remarks': ''
        }
    }
    
    current_section = None
    rows = list(ws.iter_rows(values_only=True))
    
    for i, row in enumerate(rows):
        if not any(row):
            continue
            
        first_cell = str(row[0]).strip() if row[0] else ''
        
        if first_cell == 'Reaction Conditions':
            current_section = 'reactionConditions'
        elif first_cell == 'Analytical Details':
            current_section = 'analyticalDetails'
        elif first_cell == 'Parameter' and len(row) >= 3:
            # Skip parameter header row
            continue
        elif first_cell in ['Temperature', 'Time', 'Pressure', 'Wavelength'] and current_section == 'reactionConditions':
            if len(row) >= 2 and row[1] is not None:
                key = first_cell.lower()
                settings['reactionConditions'][key] = str(row[1]).strip()
        elif first_cell in ['UPLC #', 'Method', 'Duration', 'Wavelength'] and current_section == 'analyticalDetails':
            if len(row) >= 2 and row[1] is not None:
                if first_cell == 'UPLC #':
                    settings['analyticalDetails']['uplcNumber'] = str(row[1]).strip()
                elif first_cell == 'Wavelength':
                    settings['analyticalDetails']['wavelength'] = str(row[1]).strip()
                else:
                    key = first_cell.lower()
                    settings['analyticalDetails'][key] = str(row[1]).strip()
        elif first_cell == 'Remarks':
            # Remarks content is in the next row, first cell
            if current_section and i + 1 < len(rows):
                next_row = rows[i + 1]
                if next_row and next_row[0] is not None:
                    remarks_content = str(next_row[0]).strip()
                    if remarks_content:
                        settings[current_section]['remarks'] = remarks_content
    
    return settings

def import_analytical_sheet(ws):
    """Import analytical data from Analytical Data sheet"""
    analytical_data = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Get ELN number from current experiment context for ID processing
    from state import current_experiment
    context = current_experiment.get('context', {})
    eln_number = context.get('eln', 'ELN-001')
    
    # Read analytical data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        data_item = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'Nr':
                    continue  # Skip row number
                else:
                    # Convert numeric values to proper format
                    if isinstance(value, (int, float)):
                        data_item[header] = value
                    else:
                        data_item[header] = str(value).strip()
        
        # Apply ID processing logic (same as upload functionality)
        if data_item:
            # Handle ID column mapping - if file has 'ID' column but no 'Sample ID', map it
            id_value = None
            if 'Sample ID' in data_item:
                id_value = data_item['Sample ID']
            elif 'ID' in data_item:
                # Map ID column to Sample ID
                id_value = data_item['ID']
                data_item['Sample ID'] = id_value
            
            # Process the ID/Sample ID value to ensure correct format
            if id_value and isinstance(id_value, str):
                import re
                
                # Extract well position (A1, B2, etc.) from the ID
                well_match = re.search(r'[A-H]\d{1,2}', id_value)
                if well_match:
                    well_part = well_match.group()
                    
                    # Create the correct Sample ID format: ELN_WellLocation
                    correct_sample_id = f"{eln_number}_{well_part}"
                    data_item['Sample ID'] = correct_sample_id
                    
                    print(f"Import: Mapped ID '{id_value}' to Sample ID '{correct_sample_id}'")
        
        # Only add data item if it has content beyond just the well ID
        if data_item and len([k for k, v in data_item.items() if v and str(v).strip()]) > 1:
            analytical_data.append(data_item)
    
    # Identify area columns (similar to upload functionality)
    area_columns = [col for col in headers if col.startswith('Area_')]
    
    # Create the structure expected by the frontend
    result = {
        'filename': 'imported_analytical_data.xlsx',
        'upload_date': datetime.now().isoformat(),
        'columns': headers[1:],  # Skip 'Nr' column
        'data': analytical_data,
        'shape': [len(analytical_data), len(headers) - 1] if headers else [0, 0],
        'area_columns': area_columns  # Add area columns information
    }
    
    return result

def import_results_sheet(ws):
    """Import results data from Results sheet"""
    results = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip())
    
    # Read results data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        result_item = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'nr':
                    continue  # Skip row number
                elif header == 'well':
                    result_item['well'] = str(value).strip()
                elif header == 'id':
                    result_item['id'] = str(value).strip()
                elif header == 'conversion_%':
                    result_item['conversion_percent'] = str(value).strip()
                elif header == 'yield_%':
                    result_item['yield_percent'] = str(value).strip()
                elif header == 'selectivity_%':
                    result_item['selectivity_percent'] = str(value).strip()
                else:
                    result_item[header] = str(value).strip()
        
        # Only add result item if it has a well identifier
        if result_item.get('well'):
            results.append(result_item)
    
    return results
