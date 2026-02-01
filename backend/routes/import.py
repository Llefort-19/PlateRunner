"""
Import routes blueprint.
Handles experiment data import from Excel format.
"""
import os
import tempfile
from datetime import datetime
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from state import current_experiment
from validation import (
    validate_request, validate_response,
    ExperimentContextSchema, MaterialsListSchema, ProcedureListSchema,
    ProcedureSettingsSchema, AnalyticalDataSchema, ResultsSchema
)

# Create blueprint
import_bp = Blueprint('import', __name__, url_prefix='/api/experiment')

@import_bp.route('/import', methods=['POST'])
def import_experiment():
    """Import experiment data from Excel format"""
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
        
        # Check file size (limit to 10MB)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({'error': 'File size exceeds 10MB limit'}), 400
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Load the workbook
            wb = load_workbook(tmp_path, data_only=True)
            
            # Initialize import results
            import_results = {
                'context': {'imported': False, 'data': {}},
                'materials': {'imported': False, 'count': 0, 'data': []},
                'procedure': {'imported': False, 'count': 0, 'data': []},
                'procedure_settings': {'imported': False, 'data': {}},
                'analytical_data': {'imported': False, 'count': 0, 'data': []},
                'results': {'imported': False, 'count': 0, 'data': []},
                'errors': []
            }
            
            # Import Context sheet
            if 'Context' in wb.sheetnames:
                try:
                    context_data = import_context_sheet(wb['Context'])
                    if context_data:
                        # Validate context data
                        try:
                            from validation.utils import validate_data
                            from validation.schemas import ExperimentContextSchema
                            schema = ExperimentContextSchema()
                            validated_context, errors = validate_data(
                                schema, context_data, strict_mode=False,
                                endpoint="Import Context"
                            )
                            if errors:
                                import_results['errors'].extend([f"Context validation: {err}" for err in errors])
                            context_data = validated_context
                        except Exception as validation_error:
                            import_results['errors'].append(f"Context validation error: {str(validation_error)}")
                        
                        import_results['context']['imported'] = True
                        import_results['context']['data'] = context_data
                        current_experiment['context'] = context_data
                except Exception as e:
                    import_results['errors'].append(f"Context import error: {str(e)}")
            
            # Import Materials sheet
            if 'Materials' in wb.sheetnames:
                try:
                    materials_data = import_materials_sheet(wb['Materials'])
                    if materials_data:
                        # Validate materials data
                        try:
                            from validation.utils import validate_data
                            from validation.schemas import MaterialSchema
                            schema = MaterialSchema()
                            validated_materials = []
                            for i, material in enumerate(materials_data):
                                validated_material, errors = validate_data(
                                    schema, material, strict_mode=False,
                                    endpoint=f"Import Material {i+1}"
                                )
                                if errors:
                                    import_results['errors'].extend([f"Material {i+1} validation: {err}" for err in errors])
                                validated_materials.append(validated_material)
                            materials_data = validated_materials
                        except Exception as validation_error:
                            import_results['errors'].append(f"Materials validation error: {str(validation_error)}")
                        
                        import_results['materials']['imported'] = True
                        import_results['materials']['count'] = len(materials_data)
                        import_results['materials']['data'] = materials_data
                        current_experiment['materials'] = materials_data
                except Exception as e:
                    import_results['errors'].append(f"Materials import error: {str(e)}")
            
            # Import Procedure sheet
            if 'Procedure' in wb.sheetnames:
                try:
                    procedure_data = import_procedure_sheet(wb['Procedure'])
                    if procedure_data:
                        import_results['procedure']['imported'] = True
                        import_results['procedure']['count'] = len(procedure_data)
                        import_results['procedure']['data'] = procedure_data
                        current_experiment['procedure'] = procedure_data
                except Exception as e:
                    import_results['errors'].append(f"Procedure import error: {str(e)}")
            
            # Import Procedure Settings sheet
            if 'Procedure Settings' in wb.sheetnames:
                try:
                    settings_data = import_procedure_settings_sheet(wb['Procedure Settings'])
                    if settings_data:
                        import_results['procedure_settings']['imported'] = True
                        import_results['procedure_settings']['data'] = settings_data
                        current_experiment['procedure_settings'] = settings_data
                except Exception as e:
                    import_results['errors'].append(f"Procedure Settings import error: {str(e)}")
            
            # Import Analytical data sheet
            analytical_sheet_names = [name for name in wb.sheetnames if 'Analytical data' in name]
            if analytical_sheet_names:
                try:
                    analytical_data = import_analytical_sheet(wb[analytical_sheet_names[0]])
                    if analytical_data:
                        import_results['analytical_data']['imported'] = True
                        import_results['analytical_data']['count'] = len(analytical_data)
                        import_results['analytical_data']['data'] = analytical_data
                        # Convert to new format
                        current_experiment['analytical_data'] = {
                            'selectedCompounds': [],
                            'uploadedFiles': analytical_data
                        }
                except Exception as e:
                    import_results['errors'].append(f"Analytical data import error: {str(e)}")
            
            # Import Results sheet
            results_sheet_names = [name for name in wb.sheetnames if 'Results' in name]
            if results_sheet_names:
                try:
                    results_data = import_results_sheet(wb[results_sheet_names[0]])
                    if results_data:
                        import_results['results']['imported'] = True
                        import_results['results']['count'] = len(results_data)
                        import_results['results']['data'] = results_data
                        current_experiment['results'] = results_data
                except Exception as e:
                    import_results['errors'].append(f"Results import error: {str(e)}")
            
            # Check if any data was imported
            any_imported = any([
                import_results['context']['imported'],
                import_results['materials']['imported'],
                import_results['procedure']['imported'],
                import_results['procedure_settings']['imported'],
                import_results['analytical_data']['imported'],
                import_results['results']['imported']
            ])
            
            if not any_imported:
                return jsonify({
                    'error': 'No valid experiment data found in the Excel file. Please ensure the file contains the expected sheets (Context, Materials, Procedure, etc.)'
                }), 400
            
            # Generate summary message
            summary_parts = []
            if import_results['context']['imported']:
                summary_parts.append("Context")
            if import_results['materials']['imported']:
                summary_parts.append(f"{import_results['materials']['count']} materials")
            if import_results['procedure']['imported']:
                summary_parts.append(f"{import_results['procedure']['count']} procedure entries")
            if import_results['procedure_settings']['imported']:
                summary_parts.append("Procedure settings")
            if import_results['analytical_data']['imported']:
                summary_parts.append(f"{import_results['analytical_data']['count']} analytical entries")
            if import_results['results']['imported']:
                summary_parts.append(f"{import_results['results']['count']} results")
            
            summary_message = f"Successfully imported: {', '.join(summary_parts)}"
            if import_results['errors']:
                summary_message += f". Warnings: {'; '.join(import_results['errors'])}"
            
            return jsonify({
                'message': summary_message,
                'import_results': import_results
            })
            
        finally:
            # Clean up temporary file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
                
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'}), 500

def import_context_sheet(ws):
    """Import context data from Context sheet"""
    context_data = {}
    
    # Read context data from rows
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1] is not None:
            key = str(row[0]).lower().strip()
            value = str(row[1]).strip() if row[1] is not None else ''
            
            if key == 'author':
                context_data['author'] = value
            elif key == 'date':
                # Handle date format
                if value:
                    try:
                        # Try to parse the date
                        if isinstance(row[1], datetime):
                            context_data['date'] = row[1].strftime('%Y-%m-%d')
                        else:
                            context_data['date'] = value
                    except:
                        context_data['date'] = value
            elif key == 'project':
                context_data['project'] = value
            elif key == 'eln':
                context_data['eln'] = value
            elif key == 'objective':
                context_data['objective'] = value
    
    return context_data

def import_materials_sheet(ws):
    """Import materials data from Materials sheet"""
    materials = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip())
    
    # Read materials data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        material = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'nr':
                    continue  # Skip row number
                elif header == 'chemical_name':
                    material['name'] = str(value).strip()
                elif header == 'alias':
                    material['alias'] = str(value).strip()
                elif header == 'cas_number':
                    material['cas'] = str(value).strip()
                elif header == 'molecular_weight':
                    material['molecular_weight'] = str(value).strip()
                elif header == 'smiles':
                    material['smiles'] = str(value).strip()
                elif header == 'barcode':
                    material['barcode'] = str(value).strip()
                elif header == 'role':
                    material['role'] = str(value).strip()
                elif header == 'role_id':
                    material['role_id'] = str(value).strip()
                elif header == 'source':
                    material['source'] = str(value).strip()
                elif header == 'supplier':
                    material['supplier'] = str(value).strip()
        
        # Only add material if it has a name
        if material.get('name'):
            materials.append(material)
    
    return materials

def import_procedure_sheet(ws):
    """Import procedure data from Procedure sheet"""
    procedure = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip())
    
    # Read procedure data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        well_data = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'nr':
                    continue  # Skip row number
                elif header == 'well':
                    well_data['well'] = str(value).strip()
                elif header == 'id':
                    well_data['id'] = str(value).strip()
                else:
                    # Handle compound, reagent, and solvent columns
                    well_data[header] = str(value).strip()
        
        # Only add well data if it has a well identifier
        if well_data.get('well'):
            procedure.append(well_data)
    
    return procedure

def import_procedure_settings_sheet(ws):
    """Import procedure settings from Procedure Settings sheet"""
    settings = {
        'reactionConditions': {
            'temperature': '',
            'time': '',
            'pressure': '',
            'wavelength': '',
            'remarks': ''
        },
        'analyticalDetails': {
            'uplcNumber': '',
            'method': '',
            'duration': '',
            'remarks': ''
        }
    }
    
    current_section = None
    
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
            
        first_cell = str(row[0]).strip() if row[0] else ''
        
        if first_cell == 'Reaction Conditions':
            current_section = 'reactionConditions'
        elif first_cell == 'Analytical Details':
            current_section = 'analyticalDetails'
        elif first_cell == 'Parameter' and len(row) >= 3:
            # Skip parameter header row
            continue
        elif first_cell in ['Temperature', 'Time', 'Pressure', 'Wavelength'] and current_section == 'reactionConditions':
            if len(row) >= 2 and row[1] is not None:
                key = first_cell.lower()
                settings['reactionConditions'][key] = str(row[1]).strip()
        elif first_cell in ['UPLC #', 'Method', 'Duration'] and current_section == 'analyticalDetails':
            if len(row) >= 2 and row[1] is not None:
                if first_cell == 'UPLC #':
                    settings['analyticalDetails']['uplcNumber'] = str(row[1]).strip()
                else:
                    key = first_cell.lower()
                    settings['analyticalDetails'][key] = str(row[1]).strip()
        elif first_cell == 'Remarks' and len(row) >= 2 and row[1] is not None:
            if current_section:
                settings[current_section]['remarks'] = str(row[1]).strip()
    
    return settings

def import_analytical_sheet(ws):
    """Import analytical data from Analytical data sheet"""
    analytical_data = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())  # Keep original case for header matching
    
    # Get ELN number from current experiment context for ID processing
    from state import current_experiment
    context = current_experiment.get('context', {})
    eln_number = context.get('eln', 'ELN-001')
    
    # Read analytical data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        data_item = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'Nr':
                    continue  # Skip row number
                else:
                    # Convert numeric values to proper format
                    if isinstance(value, (int, float)):
                        data_item[header] = value
                    else:
                        data_item[header] = str(value).strip()
        
        # Apply ID processing logic (same as upload functionality)
        if data_item:
            # Handle ID column mapping - if file has 'ID' column but no 'Sample ID', map it
            id_value = None
            if 'Sample ID' in data_item:
                id_value = data_item['Sample ID']
            elif 'ID' in data_item:
                # Map ID column to Sample ID
                id_value = data_item['ID']
                data_item['Sample ID'] = id_value
            
            # Process the ID/Sample ID value to ensure correct format
            if id_value and isinstance(id_value, str):
                import re
                
                # Extract well position (A1, B2, etc.) from the ID
                well_match = re.search(r'[A-H]\d{1,2}', id_value)
                if well_match:
                    well_part = well_match.group()
                    
                    # Create the correct Sample ID format: ELN_WellLocation
                    correct_sample_id = f"{eln_number}_{well_part}"
                    data_item['Sample ID'] = correct_sample_id
                    
                    print(f"Import: Mapped ID '{id_value}' to Sample ID '{correct_sample_id}'")
            
            analytical_data.append(data_item)
    
    return analytical_data

def import_results_sheet(ws):
    """Import results data from Results sheet"""
    results = []
    headers = []
    
    # Get headers from first row
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip())
    
    # Read results data
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        result_item = {}
        for i, value in enumerate(row):
            if i < len(headers) and value is not None:
                header = headers[i]
                if header == 'nr':
                    continue  # Skip row number
                elif header == 'well':
                    result_item['well'] = str(value).strip()
                elif header == 'id':
                    result_item['id'] = str(value).strip()
                elif header == 'conversion_%':
                    result_item['conversion_percent'] = str(value).strip()
                elif header == 'yield_%':
                    result_item['yield_percent'] = str(value).strip()
                elif header == 'selectivity_%':
                    result_item['selectivity_percent'] = str(value).strip()
                else:
                    result_item[header] = str(value).strip()
        
        # Only add result item if it has a well identifier
        if result_item.get('well'):
            results.append(result_item)
    
    return results
