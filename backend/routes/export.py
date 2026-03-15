"""
Export routes blueprint.
Handles experiment data export to Excel format.
All export logic lives here (moved from frontend Header.jsx).
"""
import io
import re
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from state import current_experiment

# Create blueprint
export_bp = Blueprint('export', __name__, url_prefix='/api/experiment')


@export_bp.route('/export', methods=['POST'])
def export_experiment():
    """Export experiment data to Excel format.
    
    Accepts optional JSON body with:
      - sdf_data: SDF reaction data (stored in frontend localStorage)
    Returns: .xlsx file download
    """
    try:
        # Get optional data from request body
        body = request.get_json(silent=True) or {}
        sdf_data = body.get('sdf_data', None)

        # Read experiment state
        context = current_experiment.get('context', {})
        materials = current_experiment.get('materials', [])
        procedure = current_experiment.get('procedure', [])
        procedure_settings = current_experiment.get('procedure_settings', {})
        analytical_data = current_experiment.get('analytical_data', {})
        heatmap_data = current_experiment.get('heatmap_data', {})

        # Create workbook
        wb = Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)

        plating_protocol = current_experiment.get('plating_protocol', None)

        # Build all sheets
        _build_context_sheet(wb, context, sdf_data)
        _build_materials_sheet(wb, materials)
        _build_design_sheet(wb, procedure, materials)
        _build_procedure_settings_sheet(wb, procedure_settings)
        if plating_protocol:
            _build_plating_sheet(wb, plating_protocol)
        _build_analytical_sheet(wb, analytical_data, context, procedure)
        _build_heatmap_sheets(wb, heatmap_data)
        _build_summary_sheet(wb, context, materials, procedure, analytical_data, heatmap_data, plating_protocol)

        # Generate filename
        eln = (context.get('eln') or '').strip()
        if eln:
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"{eln}_{date_str}.xlsx"
        else:
            ts = datetime.now().strftime('%Y-%m-%dT%H-%M-%S')
            filename = f"HTE_Experiment_{ts}.xlsx"

        # Write to bytes buffer and return as download
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500


@export_bp.route('/export/json', methods=['POST'])
def export_experiment_json():
    """Export experiment data as ML-ready JSON.

    Returns a JSON file with three top-level keys:
      - metadata: experiment context + procedure settings
      - materials: reference list of all materials
      - wells: flat per-well records with material amounts and analytical results
    """
    try:
        context = current_experiment.get('context', {})
        materials = current_experiment.get('materials', [])
        procedure = current_experiment.get('procedure', [])
        procedure_settings = current_experiment.get('procedure_settings', {})
        analytical_data = current_experiment.get('analytical_data', {})

        # --- metadata ---
        reaction = (procedure_settings.get('reactionConditions') or {})
        analytical_settings = (procedure_settings.get('analyticalDetails') or {})

        def _to_float(v):
            if v is None or v == '':
                return None
            try:
                return float(str(v).replace(',', '.'))
            except (ValueError, TypeError):
                return None

        metadata = {
            'author': context.get('author', ''),
            'date': context.get('date', ''),
            'project': context.get('project', ''),
            'eln': context.get('eln', ''),
            'objective': context.get('objective', ''),
            'procedure_settings': {
                'temperature': _to_float(reaction.get('temperature')),
                'time': _to_float(reaction.get('time')),
                'pressure': _to_float(reaction.get('pressure')),
                'wavelength': _to_float(reaction.get('wavelength')),
                'remarks': reaction.get('remarks', ''),
            },
            'analytical_settings': {
                'uplc_number': analytical_settings.get('uplcNumber', ''),
                'method': analytical_settings.get('method', ''),
                'duration': _to_float(analytical_settings.get('duration')),
                'remarks': analytical_settings.get('remarks', ''),
            },
            'export_version': '1.0',
            'exported_at': datetime.now().isoformat(),
        }

        # --- materials ---
        materials_list = []
        for i, mat in enumerate(materials):
            materials_list.append({
                'nr': i + 1,
                'name': mat.get('name', ''),
                'alias': mat.get('alias', ''),
                'cas': mat.get('cas', ''),
                'molecular_weight': _to_float(mat.get('molecular_weight')),
                'smiles': mat.get('smiles', ''),
                'role': mat.get('role', ''),
                'role_id': mat.get('role_id', ''),
                'barcode': mat.get('barcode', ''),
                'supplier': mat.get('supplier', ''),
                'catalog_number': mat.get('catalog_number', ''),
            })

        # --- wells ---
        # Build analytical lookup: well → { compound_name: area }
        analytical_by_well = {}
        uploaded_files = analytical_data.get('uploadedFiles', [])
        if uploaded_files:
            most_recent = uploaded_files[-1]
            data_rows = most_recent.get('data', []) if isinstance(most_recent, dict) else []
            for row in data_rows:
                well = row.get('Well', '')
                if not well:
                    continue
                compounds = {}
                name_keys = sorted([k for k in row if str(k).startswith('Name_')],
                                   key=lambda k: int(k.split('_')[1]))
                area_keys = sorted([k for k in row if str(k).startswith('Area_')],
                                   key=lambda k: int(k.split('_')[1]))
                for nk, ak in zip(name_keys, area_keys):
                    cname = row.get(nk, '')
                    carea = row.get(ak, '')
                    if cname:
                        compounds[cname] = _to_float(carea)
                analytical_by_well[well] = compounds

        wells_list = []
        for well_data in procedure:
            well = well_data.get('well', '')
            if not well:
                continue
            well_materials = well_data.get('materials', [])
            if not well_materials:
                continue

            mat_entries = []
            for mat_entry in well_materials:
                alias = mat_entry.get('alias') or mat_entry.get('name', '')
                amount = mat_entry.get('amount', '')
                unit = (mat_entry.get('unit') or '').strip()

                amount_umol = None
                amount_uL = None
                if unit in ('\u03bcL', '\u00b5L', 'uL'):
                    amount_uL = _to_float(amount)
                elif unit == 'mL':
                    v = _to_float(amount)
                    amount_uL = v * 1000 if v is not None else None
                else:
                    amount_umol = _to_float(amount)

                mat_entries.append({
                    'alias': alias,
                    'role': mat_entry.get('role', ''),
                    'amount_umol': amount_umol,
                    'amount_uL': amount_uL,
                })

            wells_list.append({
                'well': well,
                'materials': mat_entries,
                'analytical': analytical_by_well.get(well, {}),
            })

        # Sort wells naturally (A1, A2, ..., B1, ...)
        def _well_sort(w):
            wn = w['well']
            try:
                return (wn[0], int(wn[1:]))
            except (IndexError, ValueError):
                return (wn, 0)
        wells_list.sort(key=_well_sort)

        result = {
            'metadata': metadata,
            'materials': materials_list,
            'wells': wells_list,
        }

        import json
        json_str = json.dumps(result, indent=2, ensure_ascii=False)
        buf = io.BytesIO(json_str.encode('utf-8'))

        eln = (context.get('eln') or '').strip()
        date_str = datetime.now().strftime('%Y-%m-%d')
        if eln:
            filename = f"{eln}_{date_str}.json"
        else:
            filename = f"HTE_Experiment_{date_str}.json"

        return send_file(
            buf,
            mimetype='application/json',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'JSON export failed: {str(e)}'}), 500


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_context_sheet(wb, context, sdf_data):
    """Experiment Context sheet."""
    ws = wb.create_sheet('Experiment Context')

    rows = [
        ['Experiment Context'],
        [''],
        ['Author', context.get('author', '')],
        ['Date', context.get('date', '')],
        ['Project', context.get('project', '')],
        ['ELN Number', context.get('eln', '')],
        ['Objective', context.get('objective', '')],
        [''],
        ['SDF Reaction Data'],
        ['Name', 'Role', 'SMILES'],
    ]

    if sdf_data and isinstance(sdf_data, dict):
        molecules = sdf_data.get('molecules', [])
        for i, mol in enumerate(molecules):
            rows.append([
                mol.get('name') or f'ID-{str(i + 1).zfill(2)}',
                mol.get('role', ''),
                mol.get('smiles', ''),
            ])

    for row in rows:
        ws.append(row)


def _build_materials_sheet(wb, materials):
    """Materials sheet — unchanged format."""
    ws = wb.create_sheet('Materials')

    headers = [
        'Nr', 'chemical_name', 'alias', 'cas_number',
        'molecular_weight', 'smiles', 'barcode',
        'role', 'role_id', 'source', 'supplier', 'catalog_number',
    ]
    ws.append(headers)

    for i, mat in enumerate(materials):
        ws.append([
            i + 1,
            mat.get('name', ''),
            mat.get('alias', ''),
            mat.get('cas', ''),
            mat.get('molecular_weight', ''),
            mat.get('smiles', ''),
            mat.get('barcode', ''),
            mat.get('role', ''),
            mat.get('role_id', ''),
            mat.get('source', ''),
            mat.get('supplier', ''),
            mat.get('catalog_number', ''),
        ])


def _build_design_sheet(wb, procedure, materials):
    """Design sheet — long format.

    Columns: well | material_nr | material_alias | amount [µmol] | amount [µL] | dispense_order

    - One row per material per well.
    - material_nr references the Nr column in the Materials sheet (1-based).
    - Chemical reagents go into amount [µmol]; solvents into amount [µL].
    - dispense_order is the 1-based order within each well.
    - Empty wells are omitted.
    """
    ws = wb.create_sheet('Design')
    ws.append(['well', 'material_nr', 'material_alias', 'amount [µmol]', 'amount [µL]', 'dispense_order'])

    # Build lookup: (lowercase name OR lowercase alias) → Nr
    name_to_nr = {}
    alias_to_nr = {}
    for i, mat in enumerate(materials):
        nr = i + 1
        name = (mat.get('name') or '').strip().lower()
        alias = (mat.get('alias') or '').strip().lower()
        if name:
            name_to_nr[name] = nr
        if alias:
            alias_to_nr[alias] = nr

    def _resolve_nr(mat_entry):
        """Resolve material_nr from a procedure material entry."""
        name = (mat_entry.get('name') or '').strip().lower()
        alias = (mat_entry.get('alias') or '').strip().lower()
        # Try name first, then alias
        return name_to_nr.get(name) or alias_to_nr.get(alias) or ''

    # Sort procedure by well position (A1, A2, … H12)
    def _well_sort_key(entry):
        well = entry.get('well', '')
        if not well:
            return ('Z', 99)
        row_letter = well[0].upper()
        try:
            col_num = int(well[1:])
        except ValueError:
            col_num = 99
        return (row_letter, col_num)

    sorted_procedure = sorted(procedure, key=_well_sort_key)

    for well_data in sorted_procedure:
        well = well_data.get('well', '')
        if not well:
            continue
        well_materials = well_data.get('materials', [])
        if not well_materials:
            continue

        for order, mat_entry in enumerate(well_materials, start=1):
            mat_nr = _resolve_nr(mat_entry)
            display_name = mat_entry.get('alias') or mat_entry.get('name') or ''
            amount = mat_entry.get('amount', '')
            unit = (mat_entry.get('unit') or '').strip()

            # Route amount to the correct column based on unit
            if unit in ('\u03bcL', '\u00b5L', 'uL'):  # µL variants
                amount_umol = ''
                amount_ul = amount
            elif unit == 'mL':
                amount_umol = ''
                try:
                    amount_ul = float(amount) * 1000  # convert to µL
                except (ValueError, TypeError):
                    amount_ul = ''
            else:
                # Default: µmol (chemical reagent)
                amount_umol = amount
                amount_ul = ''

            ws.append([well, mat_nr, display_name, amount_umol, amount_ul, order])


def _build_procedure_settings_sheet(wb, settings):
    """Procedure Settings sheet."""
    ws = wb.create_sheet('Procedure')

    rc = settings.get('reactionConditions', {})
    ad = settings.get('analyticalDetails', {})

    rows = [
        ['Procedure Settings'],
        [''],
        ['Reaction Conditions'],
        ['Parameter', 'Value', 'Unit'],
        ['Temperature', rc.get('temperature', ''), 'degC'],
        ['Time', rc.get('time', ''), 'h'],
        ['Pressure', rc.get('pressure', ''), 'bar'],
        ['Wavelength', rc.get('wavelength', ''), 'nm'],
        [''],
        ['Remarks'],
        [rc.get('remarks', '')],
        [''],
        ['Analytical Details'],
        ['Parameter', 'Value', 'Unit'],
        ['UPLC #', ad.get('uplcNumber', ''), ''],
        ['Method', ad.get('method', ''), ''],
        ['Duration', ad.get('duration', ''), 'min'],
        ['Wavelength', ad.get('wavelength', ''), 'nm'],
        [''],
        ['Remarks'],
        [ad.get('remarks', '')],
    ]

    for row in rows:
        ws.append(row)


def _build_analytical_sheet(wb, analytical_data, context, procedure):
    """Analytical Data sheet — either uploaded data or a template."""
    ws = wb.create_sheet('Analytical Data')

    uploaded_files = analytical_data.get('uploadedFiles', [])

    if uploaded_files:
        # Use the most recent upload
        most_recent = uploaded_files[-1]
        data_rows = most_recent.get('data', []) if isinstance(most_recent, dict) else []

        if data_rows:
            # Determine columns: Well, Sample ID, then Name_X / Area_X pairs
            ordered_cols = ['Well', 'Sample ID']
            
            # Collect all unique Name_ and Area_ keys across ALL rows
            all_keys = set()
            for r in data_rows:
                all_keys.update(r.keys())
                
            name_cols = sorted(
                [k for k in all_keys if str(k).startswith('Name_')],
                key=lambda k: int(k.split('_')[1]),
            )
            area_cols = sorted(
                [k for k in all_keys if str(k).startswith('Area_')],
                key=lambda k: int(k.split('_')[1]),
            )
            max_compounds = max(len(name_cols), len(area_cols))
            for i in range(max_compounds):
                if i < len(name_cols):
                    ordered_cols.append(name_cols[i])
                if i < len(area_cols):
                    ordered_cols.append(area_cols[i])

            ws.append(ordered_cols)
            for row_data in data_rows:
                ws.append([row_data.get(c, '') for c in ordered_cols])
            return

    # Fallback: create analytical template
    _build_analytical_template(ws, analytical_data, context, procedure)


def _build_analytical_template(ws, analytical_data, context, procedure):
    """Build an empty analytical template with well IDs and compound columns."""
    eln = (context.get('eln') or 'ELN').strip()

    # Get user-selected compounds
    selected = analytical_data.get('selectedCompounds', [])
    if not selected:
        selected = ['']  # At least one empty column

    headers = ['Well', 'Sample ID']
    for i, name in enumerate(selected, start=1):
        headers.extend([f'Name_{i}', f'Area_{i}'])
    ws.append(headers)

    # Determine plate configuration
    plate_type = context.get('plate_type', '96')
    if plate_type == '24':
        rows_letters = ['A', 'B', 'C', 'D']
        cols_nums = range(1, 7)
    elif plate_type == '48':
        rows_letters = ['A', 'B', 'C', 'D', 'E', 'F']
        cols_nums = range(1, 9)
    else:
        rows_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        cols_nums = range(1, 13)

    for r in rows_letters:
        for c in cols_nums:
            well = f'{r}{c}'
            sample_id = f'{eln}_{well}'
            row = [well, sample_id]
            for name in selected:
                compound_name = name if isinstance(name, str) else ''
                row.extend([compound_name, ''])
            ws.append(row)


def _build_plating_sheet(wb, plating_protocol):
    """Plating sheet — dispensing methods, stock solutions, and order of addition.

    Handles two data formats:
    - Processed format (from ProtocolPreview export): keys 'materials', 'operations'
    - Raw format (from modal close): keys 'materialConfigs', 'dispenseOrder'
    """
    ws = wb.create_sheet('Plating')

    # Determine data format and normalize
    if 'materials' in plating_protocol:
        # Processed format from ProtocolPreview.buildProtocolData()
        materials = plating_protocol.get('materials', [])
        operations = plating_protocol.get('operations', [])
        kit_stock_entries = plating_protocol.get('kit_stock_entries', [])
        _build_plating_from_processed(ws, materials, operations, kit_stock_entries)
    elif 'materialConfigs' in plating_protocol:
        # Raw format from PlatingProtocolModal close
        material_configs = plating_protocol.get('materialConfigs', [])
        dispense_order = plating_protocol.get('dispenseOrder', [])
        _build_plating_from_raw(ws, material_configs, dispense_order)
    else:
        ws.append(['No plating protocol data available.'])
        return

    # Adjust column widths
    for col_idx in range(1, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15


def _build_plating_from_processed(ws, materials, operations, kit_stock_entries):
    """Build plating sheet from processed protocol data (ProtocolPreview export format)."""
    kit_role_ids = set(e.get('kit_id', '') for e in kit_stock_entries)

    # ── Section A: Materials Summary ──
    ws.append(['Plating Protocol — Dispensing Summary'])
    ws.append([''])
    headers = ['#', 'Material', 'Alias', 'Method', 'Solvent', 'Conc. [M]', 'Vol/Well [µL]',
               'Excess [%]', 'Total Vol [mL]', 'Mass [mg]']
    ws.append(headers)

    row_num = 0

    # Kit stock entries (grouped)
    for entry in kit_stock_entries:
        row_num += 1
        stock = entry.get('stock_solution') or {}

        conc_val = stock.get('concentration_value')
        conc_num = ''
        if conc_val is not None:
            try:
                conc_num = round(float(conc_val), 6)  # numeric, always in M
            except (ValueError, TypeError):
                pass

        tv = stock.get('total_volume_value')
        vol_num = ''
        if tv is not None:
            try:
                vol_num = round(float(tv), 4)  # numeric, in mL
            except (ValueError, TypeError):
                pass

        vpw = stock.get('amount_per_well_value')
        vpw_num = ''
        if vpw:
            try:
                vpw_f = float(vpw)
                vpw_unit = stock.get('amount_per_well_unit', '\u00b5L')
                vpw_num = round(vpw_f * 1000 if vpw_unit == 'mL' else vpw_f, 2)  # numeric, in µL
            except (ValueError, TypeError):
                pass

        excess_num = stock.get('excess', '') if stock.get('excess') is not None else ''
        member_names = entry.get('member_names', [])
        count = len(member_names)
        ws.append([
            row_num,
            f"{entry.get('kit_id', '')} ({count} materials)",
            entry.get('kit_id', ''),
            'Stock',
            stock.get('solvent_name', ''),
            conc_num,
            vpw_num,
            excess_num,
            vol_num,
            '\u2014'  # em-dash: mass varies per member
        ])

    # Non-kit materials
    for mat in materials:
        role_id = mat.get('role_id', '')
        # Skip individual kit-stock members (already represented by kit entry)
        if role_id in kit_role_ids and mat.get('dispensing_method') == 'stock':
            continue
        # Skip kit neat members too (represented by kit entry)
        if role_id in kit_role_ids:
            continue

        row_num += 1
        method = mat.get('dispensing_method', 'neat')
        total_unit = mat.get('total_amount_unit', 'μmol')
        is_solvent = total_unit in ('μL', 'mL')

        if method == 'stock':
            method_label = 'Stock'
        elif is_solvent:
            method_label = 'Solvent'
        else:
            method_label = 'Neat'

        stock = mat.get('stock_solution') or {}
        conc_str = ''
        vol_str = ''
        vpw_str = ''
        excess_str = ''
        mass_str = ''

        if method == 'stock' and stock:
            conc_val = stock.get('concentration_value')
            if conc_val is not None:
                try:
                    conc_str = round(float(conc_val), 6)  # numeric, always in M
                except (ValueError, TypeError):
                    pass
            tv = stock.get('total_volume_value')
            if tv is not None:
                try:
                    vol_str = round(float(tv), 4)  # numeric, in mL
                except (ValueError, TypeError):
                    pass
            vpw = stock.get('amount_per_well_value')
            if vpw:
                try:
                    vpw_f = float(vpw)
                    vpw_unit = stock.get('amount_per_well_unit', '\u00b5L')
                    vpw_str = round(vpw_f * 1000 if vpw_unit == 'mL' else vpw_f, 2)  # numeric, in µL
                except (ValueError, TypeError):
                    pass
            excess_str = stock.get('excess', '') if stock.get('excess') is not None else ''
            mass_val = mat.get('calculated_mass_value')
            if mass_val is not None:
                try:
                    mass_str = f"{float(mass_val):.2f}"
                except (ValueError, TypeError):
                    pass

        # If cocktail, write a header row and then the components
        if mat.get('is_cocktail') or mat.get('isCocktail'):
            components = mat.get('components', [])
            comp_names = [c.get('alias') or c.get('name', '') for c in components]
            ws.append([
                row_num,
                f"{mat.get('alias', '') or mat.get('name', '')} (Premixed)",
                mat.get('alias', '') or mat.get('name', ''),
                'Stock',
                stock.get('solvent_name', '') if stock else '',
                '',
                '',
                stock.get('excess', '') if stock.get('excess') is not None else '',
                vol_str,
                ''
            ])
            for comp in components:
                c_stock = comp.get('stockSolution') or {}

                # Prefer pre-calculated concentration from buildProtocolData (accurate, uses cocktail's amountPerWell)
                c_conc_val = comp.get('calculated_concentration_value')
                if c_conc_val is not None:
                    try:
                        c_conc_str = round(float(c_conc_val), 6)  # numeric, always in M
                    except (ValueError, TypeError):
                        c_conc_str = ''
                else:
                    # Fallback to stored concentration (may be stale from individual config)
                    c_conc = c_stock.get('concentration') or {}
                    if c_conc.get('value'):
                        try:
                            cv = float(c_conc['value'])
                            if c_conc.get('unit') == 'mM':
                                cv /= 1000  # normalise to M
                            c_conc_str = round(cv, 6)
                        except (ValueError, TypeError):
                            c_conc_str = ''
                    else:
                        c_conc_str = ''

                # Prefer top-level calculated_mass_value, then legacy calculatedMass fields
                comp_mass_val = comp.get('calculated_mass_value')
                if comp_mass_val is None:
                    comp_mass_val = (comp.get('calculatedMass') or {}).get('value')
                if comp_mass_val is None:
                    comp_mass_val = (c_stock.get('calculatedMass') or {}).get('value')
                try:
                    comp_mass_str = f"{float(comp_mass_val):.2f}" if comp_mass_val is not None else ''
                except (ValueError, TypeError):
                    comp_mass_str = ''

                ws.append([
                    '',
                    f"  ↳ {comp.get('alias') or comp.get('name', '')}",
                    comp.get('alias') or comp.get('name', ''),
                    '',
                    c_stock.get('solvent', {}).get('name', '') if c_stock else '',
                    c_conc_str,
                    '',
                    '',
                    '',
                    comp_mass_str
                ])
        else:
            ws.append([
                row_num,
                mat.get('name', ''),
                mat.get('alias', ''),
                method_label,
                stock.get('solvent_name', '') if stock else '',
                conc_str,
                vpw_str,
                excess_str,
                vol_str,
                mass_str
            ])

    # ── Section B: Order of Addition ──
    ws.append([''])
    ws.append(['Order of Addition'])
    ws.append(['Step', 'Type', 'Description'])

    step_num = 0
    for op in operations:
        op_type = op.get('type', '')
        # Guard against stale dispense ops referencing a materialIndex that no longer exists
        if op_type == 'dispense':
            mat_idx = op.get('materialIndex', 0)
            if mat_idx >= len(materials):
                continue
        step_num += 1
        description = _format_plating_operation(op, materials)
        type_label = {
            'dispense': 'Dispense',
            'kit': 'Kit',
            'wait': 'Wait',
            'stir': 'Stir',
            'evaporate': 'Evaporate',
            'note': 'Note'
        }.get(op_type, op_type)
        ws.append([step_num, type_label, description])


def _build_plating_from_raw(ws, material_configs, dispense_order):
    """Build plating sheet from raw modal data (materialConfigs + dispenseOrder)."""
    # ── Section A: Materials Summary ──
    ws.append(['Plating Protocol — Dispensing Summary'])
    ws.append([''])
    headers = ['#', 'Material', 'Alias', 'Method', 'Solvent', 'Conc. [M]', 'Vol/Well [µL]',
               'Excess [%]', 'Total Vol [mL]', 'Mass [mg]']
    ws.append(headers)

    # Group kit materials
    kit_groups = {}
    regular_materials = []
    for idx, mat in enumerate(material_configs):
        role_id = mat.get('role_id', '')
        if role_id and role_id.startswith('kit_'):
            if role_id not in kit_groups:
                kit_groups[role_id] = []
            kit_groups[role_id].append((idx, mat))
        else:
            regular_materials.append((idx, mat))

    row_num = 0

    # Kit groups (one row per kit)
    for kit_id, members in sorted(kit_groups.items()):
        row_num += 1
        # Use first member for shared stock details
        ref_mat = members[0][1]
        method = ref_mat.get('dispensingMethod', 'neat')
        stock = ref_mat.get('stockSolution') or {}

        conc_str = ''
        vol_str = ''
        vpw_str = ''
        excess_str = ''
        solvent_name = ''

        if method == 'stock' and stock:
            solvent = stock.get('solvent') or {}
            solvent_name = solvent.get('name', '')
            conc = stock.get('concentration') or {}
            if conc.get('value'):
                try:
                    cv = float(conc['value'])
                    if conc.get('unit') == 'mM':
                        cv /= 1000  # normalise to M
                    conc_str = round(cv, 6)  # numeric, in M
                except (ValueError, TypeError):
                    pass
            apw = stock.get('amountPerWell') or {}
            if apw.get('value'):
                try:
                    apw_f = float(apw['value'])
                    apw_unit = apw.get('unit', '\u00b5L')
                    vpw_str = round(apw_f * 1000 if apw_unit == 'mL' else apw_f, 2)  # numeric, in µL
                except (ValueError, TypeError):
                    pass
            excess = stock.get('excess')
            if excess is not None:
                excess_str = excess  # numeric, no %
            tv = stock.get('totalVolume') or {}
            if tv.get('value'):
                try:
                    tv_f = float(tv['value'])
                    tv_unit = tv.get('unit', 'mL')
                    if tv_unit in ('\u03bcL', '\u00b5L', 'uL'):
                        tv_f /= 1000  # convert µL → mL
                    vol_str = round(tv_f, 4)  # numeric, in mL
                except (ValueError, TypeError):
                    pass

        method_label = 'Stock' if method == 'stock' else 'Neat'
        ws.append([
            row_num,
            f"{kit_id} ({len(members)} materials)",
            kit_id,
            method_label,
            solvent_name,
            conc_str,
            vpw_str,
            excess_str,
            vol_str,
            '\u2014'
        ])

    # Regular materials
    for orig_idx, mat in regular_materials:
        row_num += 1
        method = mat.get('dispensingMethod', 'neat')
        total_unit = (mat.get('totalAmount') or {}).get('unit', 'μmol')
        is_solvent = total_unit in ('μL', 'mL')

        if method == 'stock':
            method_label = 'Stock'
        elif is_solvent:
            method_label = 'Solvent'
        else:
            method_label = 'Neat'

        stock = mat.get('stockSolution') or {}
        conc_str = ''
        vol_str = ''
        vpw_str = ''
        excess_str = ''
        mass_str = ''
        solvent_name = ''

        if method == 'stock' and stock:
            solvent = stock.get('solvent') or {}
            solvent_name = solvent.get('name', '')
            conc = stock.get('concentration') or {}
            if conc.get('value'):
                try:
                    cv = float(conc['value'])
                    if conc.get('unit') == 'mM':
                        cv /= 1000  # normalise to M
                    conc_str = round(cv, 6)  # numeric, in M
                except (ValueError, TypeError):
                    pass
            apw = stock.get('amountPerWell') or {}
            if apw.get('value'):
                try:
                    apw_f = float(apw['value'])
                    apw_unit = apw.get('unit', '\u00b5L')
                    vpw_str = round(apw_f * 1000 if apw_unit == 'mL' else apw_f, 2)  # numeric, in µL
                except (ValueError, TypeError):
                    pass
            excess = stock.get('excess')
            if excess is not None:
                excess_str = excess  # numeric, no %
            tv = stock.get('totalVolume') or {}
            if tv.get('value'):
                try:
                    tv_f = float(tv['value'])
                    tv_unit = tv.get('unit', 'mL')
                    if tv_unit in ('\u03bcL', '\u00b5L', 'uL'):
                        tv_f /= 1000  # convert µL → mL
                    vol_str = round(tv_f, 4)  # numeric, in mL
                except (ValueError, TypeError):
                    pass
            calc_mass = mat.get('calculatedMass') or {}
            if calc_mass.get('value'):
                try:
                    mass_str = f"{float(calc_mass['value']):.2f}"
                except (ValueError, TypeError):
                    pass
        elif method == 'neat' and not is_solvent:
            # Neat material: total mass = totalAmount (μmol) × MW (g/mol) / 1000 → mg
            mw = mat.get('molecular_weight')
            total_amt = (mat.get('totalAmount') or {}).get('value')
            if mw and total_amt:
                try:
                    mass_str = f"{float(total_amt) * float(mw) / 1000:.2f}"
                except (ValueError, TypeError):
                    pass

        if mat.get('isCocktail'):
            components = mat.get('components', [])
            # Pre-calculate cocktail's amountPerWell in μL for component concentration
            cocktail_apw = (stock.get('amountPerWell') or {}) if stock else {}
            cocktail_vpw = cocktail_apw.get('value')
            cocktail_vpw_unit = cocktail_apw.get('unit', 'μL')
            try:
                cocktail_vpw_ul = float(cocktail_vpw) * 1000 if cocktail_vpw_unit == 'mL' else float(cocktail_vpw) if cocktail_vpw else None
            except (ValueError, TypeError):
                cocktail_vpw_ul = None

            cocktail_excess = stock.get('excess') or 0 if stock else 0

            ws.append([
                row_num,
                f"{mat.get('alias', mat.get('name', ''))} (Premixed)",
                mat.get('alias', mat.get('name', '')),
                'Stock',
                solvent_name,
                '',
                '',
                cocktail_excess if cocktail_excess else '',  # numeric, no %
                vol_str,
                ''
            ])
            for comp in components:
                c_stock = comp.get('stockSolution') or {}
                c_solvent = c_stock.get('solvent') or {}

                # Calculate concentration from component's wellAmounts + cocktail's amountPerWell
                c_conc_num = ''
                c_well_amounts = comp.get('wellAmounts') or {}
                if cocktail_vpw_ul and c_well_amounts:
                    try:
                        vals = [float(v.get('value', 0)) for v in c_well_amounts.values() if v.get('value')]
                        if vals:
                            c_conc_num = round(min(vals) / cocktail_vpw_ul, 6)  # numeric, in M
                    except (ValueError, TypeError):
                        pass
                if not c_conc_num:
                    # Fallback to stored concentration, normalised to M
                    c_conc = c_stock.get('concentration') or {}
                    if c_conc.get('value'):
                        try:
                            cv = float(c_conc['value'])
                            if c_conc.get('unit') == 'mM':
                                cv /= 1000
                            c_conc_num = round(cv, 6)
                        except (ValueError, TypeError):
                            pass

                # Mass: check stockSolution.calculatedMass first (set by batch-apply),
                # then top-level calculatedMass, then calculate from totalAmount × MW
                comp_mass_val = None
                comp_mass_dict = c_stock.get('calculatedMass') or comp.get('calculatedMass') or {}
                comp_mass_val = comp_mass_dict.get('value')
                if not comp_mass_val:
                    # Fallback: calculate from totalAmount × MW × (1 + excess%)
                    comp_mw = comp.get('molecular_weight')
                    comp_total = (comp.get('totalAmount') or {}).get('value')
                    if comp_mw and comp_total:
                        try:
                            comp_mass_val = float(comp_total) * float(comp_mw) * (1 + cocktail_excess / 100) / 1000
                        except (ValueError, TypeError):
                            pass
                try:
                    comp_mass_str = f"{float(comp_mass_val):.2f}" if comp_mass_val is not None else ''
                except (ValueError, TypeError):
                    comp_mass_str = ''

                ws.append([
                    '',
                    f"  ↳ {comp.get('alias') or comp.get('name', '')}",
                    comp.get('alias') or comp.get('name', ''),
                    '',
                    c_solvent.get('name', ''),
                    c_conc_num,
                    '',
                    '',
                    '',
                    comp_mass_str
                ])
        else:
            ws.append([
                row_num,
                mat.get('name', ''),
                mat.get('alias', mat.get('name', '')),
                method_label,
                solvent_name,
                conc_str,
                vpw_str,
                excess_str,
                vol_str,
                mass_str
            ])

    # ── Section B: Order of Addition ──
    ws.append([''])
    ws.append(['Order of Addition'])
    ws.append(['Step', 'Type', 'Description'])

    step_num = 0
    for op in dispense_order:
        op_type = op.get('type', '')
        # Skip stale dispense ops that point to a material index no longer in materialConfigs
        # (can happen when two materials were combined into a cocktail without the dispenseOrder
        # being updated — e.g. state saved before the frontend fix was applied).
        if op_type == 'dispense':
            mat_idx = op.get('materialIndex', 0)
            if mat_idx >= len(material_configs):
                continue
        step_num += 1
        description = _format_plating_operation_raw(op, material_configs)
        type_label = {
            'dispense': 'Dispense',
            'kit': 'Kit',
            'wait': 'Wait',
            'stir': 'Stir',
            'evaporate': 'Evaporate',
            'note': 'Note'
        }.get(op_type, op_type)
        ws.append([step_num, type_label, description])


def _format_plating_operation(op, materials):
    """Format a single operation for the plating order of addition (processed format)."""
    op_type = op.get('type', '')
    if op_type == 'dispense':
        mat_idx = op.get('materialIndex', 0)
        if mat_idx < len(materials):
            mat = materials[mat_idx]
            name = mat.get('alias') or mat.get('name', 'Material')
            method = mat.get('dispensing_method', 'neat')
            total_unit = mat.get('total_amount_unit', 'μmol')
            is_solvent = total_unit in ('μL', 'mL')
            method_label = 'Solvent' if is_solvent else ('Stock' if method == 'stock' else 'Neat')
            well_count = len(mat.get('well_amounts', {}))
            return f"{name} — {method_label} • {well_count} wells"
        return 'Unknown material'
    elif op_type == 'kit':
        kit_id = op.get('kitId', 'Kit')
        count = len(op.get('materialIndices', []))
        text = f"{kit_id} — {count} materials"
        note = op.get('note', '')
        if note:
            text += f" — {note}"
        return text
    elif op_type == 'wait':
        return f"Wait {op.get('duration', '—')} {op.get('unit', 'min')}"
    elif op_type == 'stir':
        text = f"Stir at {op.get('temperature', '—')}°C for {op.get('duration', '—')} {op.get('unit', 'min')}"
        if op.get('rpm'):
            text += f" @ {op['rpm']} RPM"
        return text
    elif op_type == 'evaporate':
        return 'Evaporate solvents'
    elif op_type == 'note':
        return op.get('text', 'Note')
    return 'Unknown operation'


def _format_plating_operation_raw(op, material_configs):
    """Format a single operation for the plating order of addition (raw format)."""
    op_type = op.get('type', '')
    if op_type == 'dispense':
        mat_idx = op.get('materialIndex', 0)
        if mat_idx < len(material_configs):
            mat = material_configs[mat_idx]
            name = mat.get('alias') or mat.get('name', 'Material')
            method = mat.get('dispensingMethod', 'neat')
            total_unit = (mat.get('totalAmount') or {}).get('unit', 'μmol')
            is_solvent = total_unit in ('μL', 'mL')
            method_label = 'Solvent' if is_solvent else ('Stock' if method == 'stock' else 'Neat')
            well_count = len(mat.get('wellAmounts', {}))
            return f"{name} — {method_label} • {well_count} wells"
        return 'Unknown material'
    elif op_type == 'kit':
        kit_id = op.get('kitId', 'Kit')
        count = len(op.get('materialIndices', []))
        text = f"{kit_id} — {count} materials"
        note = op.get('note', '')
        if note:
            text += f" — {note}"
        return text
    elif op_type == 'wait':
        return f"Wait {op.get('duration', '—')} {op.get('unit', 'min')}"
    elif op_type == 'stir':
        text = f"Stir at {op.get('temperature', '—')}°C for {op.get('duration', '—')} {op.get('unit', 'min')}"
        if op.get('rpm'):
            text += f" @ {op['rpm']} RPM"
        return text
    elif op_type == 'evaporate':
        return 'Evaporate solvents'
    elif op_type == 'note':
        return op.get('text', 'Note')
    return 'Unknown operation'


def _heatmap_cell_color(value, min_val, max_val, color_scheme='blue'):
    """Compute a fill hex color (RRGGBB) for a heatmap cell.

    Mirrors the frontend getHeatmapColor() function exactly so that the Excel
    export matches what the user sees in the app.
    """
    try:
        value = float(value)
        min_val = float(min_val)
        max_val = float(max_val)
    except (TypeError, ValueError):
        return None

    if value == 0 or min_val == max_val:
        return None  # leave cell uncoloured (matches frontend empty-well treatment)

    normalized = max(0.0, min(1.0, (value - min_val) / (max_val - min_val)))

    def lerp(a, b, t):
        return round(a + (b - a) * t)

    def rgb(r, g, b):
        return f'FF{int(r):02X}{int(g):02X}{int(b):02X}'

    if color_scheme == 'blue-yellow-red':
        # #2c7bb6 → #abd9e9 → #ffffbf → #fdae61 → #d7191c
        if normalized <= 0.25:
            t = normalized / 0.25
            return rgb(lerp(44, 171, t), lerp(123, 217, t), lerp(182, 233, t))
        elif normalized <= 0.5:
            t = (normalized - 0.25) / 0.25
            return rgb(lerp(171, 255, t), lerp(217, 255, t), lerp(233, 191, t))
        elif normalized <= 0.75:
            t = (normalized - 0.5) / 0.25
            return rgb(lerp(255, 253, t), lerp(255, 174, t), lerp(191, 97, t))
        else:
            t = (normalized - 0.75) / 0.25
            return rgb(lerp(253, 215, t), lerp(174, 25, t), lerp(97, 28, t))

    elif color_scheme == 'green-blue':
        # #ffffd9 → #edf8b1 → #c7e9b4 → #7fcdbb → #41b6c4
        if normalized <= 0.25:
            t = normalized / 0.25
            return rgb(lerp(255, 237, t), lerp(255, 248, t), lerp(217, 177, t))
        elif normalized <= 0.5:
            t = (normalized - 0.25) / 0.25
            return rgb(lerp(237, 199, t), lerp(248, 233, t), lerp(177, 180, t))
        elif normalized <= 0.75:
            t = (normalized - 0.5) / 0.25
            return rgb(lerp(199, 127, t), lerp(233, 205, t), lerp(180, 187, t))
        else:
            t = (normalized - 0.75) / 0.25
            return rgb(lerp(127, 65, t), lerp(205, 182, t), lerp(187, 196, t))

    elif color_scheme == 'purple-green-yellow':
        # #440154 → #3b528b → #21918c → #5ec962 → #fde725
        if normalized <= 0.25:
            t = normalized / 0.25
            return rgb(lerp(68, 59, t), lerp(1, 82, t), lerp(84, 139, t))
        elif normalized <= 0.5:
            t = (normalized - 0.25) / 0.25
            return rgb(lerp(59, 33, t), lerp(82, 145, t), lerp(139, 140, t))
        elif normalized <= 0.75:
            t = (normalized - 0.5) / 0.25
            return rgb(lerp(33, 94, t), lerp(145, 201, t), lerp(140, 98, t))
        else:
            t = (normalized - 0.75) / 0.25
            return rgb(lerp(94, 253, t), lerp(201, 231, t), lerp(98, 37, t))

    else:  # 'blue' (default)
        intensity = int(normalized * 255)
        v = 255 - intensity
        return rgb(v, v, 255)


def _font_color_for_bg(hex_color):
    """Return black or white font color (FFRRGGBB) for legibility on hex_color."""
    r = int(hex_color[2:4], 16)
    g = int(hex_color[4:6], 16)
    b = int(hex_color[6:8], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return 'FF000000' if luminance > 0.5 else 'FFFFFFFF'


def _build_heatmap_sheets(wb, heatmap_data):
    """Heatmap sheets (one per heatmap)."""
    if not heatmap_data:
        heatmap_data = {}

    heatmaps = heatmap_data.get('heatmaps', [])

    if heatmaps:
        for idx, hm in enumerate(heatmaps):
            sheet_name = f'Heatmap_{idx + 1}'
            ws = wb.create_sheet(sheet_name)

            rows = [
                [f"Heatmap {idx + 1}: {hm.get('title', 'Untitled')}"],
                [''],
                ['Formula:', hm.get('formula', 'No formula')],
                ['Color Scheme:', hm.get('colorScheme', 'blue')],
                ['Min Value:', hm.get('min', 0)],
                ['Max Value:', hm.get('max', 0)],
                [''],
            ]
            for row in rows:
                ws.append(row)

            col_labels = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
            row_labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

            ws.append([''] + col_labels)

            hm_data = hm.get('data', [])
            heatmap_color_schemes = heatmap_data.get('heatmapColorSchemes', {})
            heatmap_id = hm.get('id')
            color_scheme = heatmap_color_schemes.get(str(heatmap_id), 'blue')
            min_val = hm.get('min', 0)
            max_val = hm.get('max', 0)


            for ri, data_row in enumerate(hm_data):
                label = row_labels[ri] if ri < len(row_labels) else str(ri + 1)
                row_num = ws.max_row + 1
                ws.cell(row=row_num, column=1, value=label)
                for ci, value in enumerate(data_row):
                    cell = ws.cell(row=row_num, column=ci + 2, value=round(value, 3) if value else '')
                    if value:
                        hex_color = _heatmap_cell_color(value, min_val, max_val, color_scheme)
                        if hex_color:
                            cell.fill = PatternFill(
                                start_color=hex_color,
                                end_color=hex_color,
                                fill_type='solid'
                            )
                            cell.font = Font(color=_font_color_for_bg(hex_color))
    else:
        ws = wb.create_sheet('Heatmap Data')
        ws.append(['Heatmap Data'])
        ws.append([''])
        ws.append(['No heatmaps generated yet.'])


def _build_summary_sheet(wb, context, materials, procedure, analytical_data, heatmap_data, plating_protocol=None):
    """Summary overview sheet."""
    ws = wb.create_sheet('Summary')

    heatmaps = (heatmap_data or {}).get('heatmaps', [])
    uploaded = analytical_data.get('uploadedFiles', [])
    has_plating = plating_protocol is not None

    rows = [
        ['HTE Experiment Summary'],
        [''],
        ['Experiment Information'],
        ['Author:', context.get('author', 'Not specified')],
        ['Date:', context.get('date', 'Not specified')],
        ['Project:', context.get('project', 'Not specified')],
        ['ELN Number:', context.get('eln', 'Not specified')],
        ['Objective:', context.get('objective', 'Not specified')],
        [''],
        ['Data Summary'],
        ['Materials Count:', len(materials)],
        ['Wells with Data:', sum(1 for w in procedure if w.get('materials'))],
        ['Plating Protocol:', 'Configured' if has_plating else 'Not configured'],
        ['Analytical Data Files:', len(uploaded)],
        ['Heatmaps Generated:', len(heatmaps)],
        [''],
        ['Sheet Contents'],
        ['1. Experiment Context - Basic experiment information and SDF reaction data'],
        ['2. Materials - All chemical materials with properties and roles'],
        ['3. Design - Well contents in long format (well, material_nr, alias, amount [µmol], amount [µL], dispense_order)'],
        ['4. Procedure - Reaction conditions and analytical details'],
    ]

    sheet_num = 5
    if has_plating:
        rows.append([f'{sheet_num}. Plating - Dispensing methods, stock solutions, and order of addition'])
        sheet_num += 1
    rows.append([f'{sheet_num}. Analytical Data - Uploaded analytical results or template'])
    sheet_num += 1
    rows.append([f'{sheet_num}. Heatmap - Generated heatmaps with formulas and color schemes'])
    sheet_num += 1
    rows.append([f'{sheet_num}. Summary - This overview sheet'])

    rows.extend([
        [''],
        ['Export Information'],
        ['Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Export Version:', '2.0'],
        ['File Format:', 'Excel (.xlsx)'],
    ])

    for row in rows:
        ws.append(row)
