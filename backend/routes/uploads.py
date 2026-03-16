"""
Upload routes blueprint.
Handles file uploads for analytical data and materials.
"""
import os
import pandas as pd
from datetime import datetime
from flask import Blueprint, request, jsonify
from state import current_experiment, inventory_data, load_inventory
from config import get_config
import logging

logger = logging.getLogger(__name__)

# Create blueprint
uploads_bp = Blueprint('uploads', __name__, url_prefix='/api/experiment')

@uploads_bp.route('/analytical/template', methods=['POST'])
def export_analytical_template():
    """Export analytical data template"""
    try:
        data = request.get_json()
        compounds = data.get('compounds', [])
        
        if not compounds:
            return jsonify({'error': 'No compounds provided'}), 400
        
        # Create a new workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytical Data"
        
        # Add headers - Well, Sample ID, then compound columns
        headers = ['Well', 'Sample ID']
        for i, compound in enumerate(compounds, 1):
            headers.extend([f'Name_{i}', f'Area_{i}'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Get experiment context for ELN number and plate type
        context = current_experiment.get('context', {})
        eln_number = context.get('eln', 'ELN-001')  # Use 'eln' field from context
        plate_type = context.get('plate_type', '96')  # Default to 96-well if not set
        
        # Generate wells based on plate type
        def get_plate_config(plate_type):
            if plate_type == "24":
                return {
                    'rows': ["A", "B", "C", "D"],
                    'columns': ["1", "2", "3", "4", "5", "6"]
                }
            elif plate_type == "48":
                return {
                    'rows': ["A", "B", "C", "D", "E", "F"],
                    'columns': ["1", "2", "3", "4", "5", "6", "7", "8"]
                }
            else:  # Default to 96-well
                return {
                    'rows': ["A", "B", "C", "D", "E", "F", "G", "H"],
                    'columns': ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
                }
        
        plate_config = get_plate_config(plate_type)
        rows = plate_config['rows']
        columns = plate_config['columns']
        
        row_num = 2
        for row in rows:
            for col in columns:
                well = f"{row}{col}"
                sample_id = f"{eln_number}_{well}"  # Use underscore instead of hyphen
                
                # Add well and sample ID
                ws.cell(row=row_num, column=1, value=well)
                ws.cell(row=row_num, column=2, value=sample_id)
                
                # Add compound name and area columns
                for compound_idx, compound in enumerate(compounds):
                    # Cpd_(number)_Name column
                    ws.cell(row=row_num, column=3 + (compound_idx * 2), value=compound)
                    # Cmpd_(number)_area column (empty for user to fill)
                    ws.cell(row=row_num, column=4 + (compound_idx * 2), value="")
                
                row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Read the file and return it
        with open(tmp_file_path, 'rb') as f:
            file_content = f.read()
        
        # Clean up temporary file
        os.unlink(tmp_file_path)
        
        from flask import send_file
        import io
        return send_file(
            io.BytesIO(file_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Analytical_Template_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@uploads_bp.route('/analytical/upload', methods=['POST'])
def upload_analytical_data():
    """Upload analytical data file"""
    try:
        logger.debug("Upload endpoint called")
        
        if 'file' not in request.files:
            logger.debug("No file in request.files")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        logger.debug(f"File received: {file.filename}")
        
        # Validate file upload with security checks
        try:
            from security.file_validation import validate_file_upload
            is_valid, error_msg, safe_filename = validate_file_upload(file)
            if not is_valid:
                return jsonify({'error': error_msg}), 400
        except ImportError:
            # Fallback to basic validation if security module not available
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            allowed_extensions = {'.xlsx', '.xls', '.csv'}
            file_ext = os.path.splitext(file.filename)[1].lower()
            if file_ext not in allowed_extensions:
                return jsonify({'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'}), 400
            safe_filename = file.filename
        
        # Get file extension for later use
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        # Read the file based on its type
        try:
            logger.debug("Attempting to read file")
            if file_ext == '.csv':
                df = pd.read_csv(file)
            else:  # Excel files
                # For multi-sheet Excel files, find the best sheet for analytical data
                # rather than blindly reading the first sheet (which may be 'Experiment Context')
                excel_file = pd.ExcelFile(file)
                sheet_names = excel_file.sheet_names
                target_sheet = None

                if len(sheet_names) == 1:
                    # Single-sheet file: use the only sheet
                    target_sheet = sheet_names[0]
                else:
                    # Multi-sheet file: look for an analytical data sheet by name
                    analytical_keywords = ['analytical', 'results', 'data', 'analysis']
                    for sheet in sheet_names:
                        if any(kw in sheet.lower() for kw in analytical_keywords):
                            target_sheet = sheet
                            break

                    # If no match by name, find the sheet with Area_* columns
                    if target_sheet is None:
                        for sheet in sheet_names:
                            temp_df = pd.read_excel(excel_file, sheet_name=sheet, nrows=0)
                            area_cols = [c for c in temp_df.columns if c.startswith('Area_')]
                            if area_cols:
                                target_sheet = sheet
                                break

                    # Last resort: use the first sheet
                    if target_sheet is None:
                        target_sheet = sheet_names[0]

                logger.debug(f"Excel sheets: {sheet_names}, selected: '{target_sheet}'")
                df = pd.read_excel(excel_file, sheet_name=target_sheet)
            logger.debug(f"File read successfully. Shape: {df.shape}")
        except Exception as e:
            logger.error(f"Error reading file: {str(e)}")
            return jsonify({'error': f'Error reading file: {str(e)}'}), 400
        
        # Basic validation - check if file has expected structure
        if len(df.columns) < 2:
            logger.debug(f"File has insufficient columns: {len(df.columns)}")
            return jsonify({'error': 'File must have at least 2 columns (Well and Sample ID)'}), 400
        
        # Validate area columns (columns containing "Area_" in their names)
        area_columns = [col for col in df.columns if col.startswith('Area_')]
        logger.debug(f"Found area columns: {area_columns}")
        
        # Pre-process area columns: replace empty cells with 0
        for col in area_columns:
            # First, replace various empty representations with NaN
            df[col] = df[col].replace(['', ' ', 'nan', 'NaN', 'None', None], pd.NA)
            # Fill NaN values with 0
            df[col] = df[col].fillna(0)
            logger.debug(f"Pre-processed column {col}: replaced empty cells with 0")
        
        # Check if area columns contain only numerical data
        invalid_area_columns = []
        for col in area_columns:
            try:
                # Convert to numeric, coercing errors to NaN
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                # Check if there are any NaN values (indicating non-numeric data)
                # After our pre-processing, NaN should only occur for truly non-numeric data
                if numeric_col.isna().any():
                    invalid_area_columns.append(col)
                    logger.debug(f"Column {col} contains non-numeric data")
                    # Show which values couldn't be converted
                    nan_mask = numeric_col.isna()
                    problematic_values = df.loc[nan_mask, col].unique()
                    logger.debug(f"Problematic values in {col}: {problematic_values}")
            except Exception as e:
                invalid_area_columns.append(col)
                logger.error(f"Error validating column {col}: {str(e)}")
        
        if invalid_area_columns:
            return jsonify({
                'error': f'Area columns must contain only numerical data. Invalid columns: {", ".join(invalid_area_columns)}'
            }), 400
        
        logger.debug(f"Current experiment state before upload: {list(current_experiment.keys())}")
        
        # Process the data to ensure correct ID format
        # Get ELN number from context
        context = current_experiment.get('context', {})
        eln_number = context.get('eln', 'ELN-001')

        # Drop completely empty rows (all NaN) before processing
        df = df.dropna(how='all')

        # Process each row to ensure correct ID format
        import re
        import math
        processed_data = []
        for _, row in df.iterrows():
            raw_row = row.to_dict()  # Convert pandas Series to dict

            # Sanitize all values: replace NaN/NaT with None for JSON compatibility
            # Python's json.dumps outputs NaN as bare 'NaN' which is invalid JSON
            # and causes JavaScript JSON.parse() to fail silently
            processed_row = {}
            for k, v in raw_row.items():
                if v is None:
                    processed_row[k] = None
                elif isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                    processed_row[k] = None
                elif hasattr(v, 'item'):
                    # Convert numpy scalar types to native Python types for JSON serialization
                    processed_row[k] = v.item()
                else:
                    processed_row[k] = v

            # Skip rows where both Well and Sample ID are empty (trailing blank rows)
            well_val = processed_row.get('Well')
            sample_id_val = processed_row.get('Sample ID') or processed_row.get('ID')
            if well_val is None and sample_id_val is None:
                continue

            # Handle ID column mapping - if file has 'ID' column but no 'Sample ID', map it
            id_value = None
            if 'Sample ID' in processed_row:
                id_value = processed_row['Sample ID']
            elif 'ID' in processed_row:
                # Map ID column to Sample ID
                id_value = processed_row['ID']
                processed_row['Sample ID'] = id_value

            # Process the ID/Sample ID value to ensure correct format
            if id_value and isinstance(id_value, str):
                # Extract well position (A1, B2, etc.) from the ID
                well_match = re.search(r'[A-H]\d{1,2}', id_value)
                if well_match:
                    well_part = well_match.group()

                    # Create the correct Sample ID format: ELN_WellLocation
                    correct_sample_id = f"{eln_number}_{well_part}"
                    processed_row['Sample ID'] = correct_sample_id

                    logger.debug(f"Mapped ID '{id_value}' to Sample ID '{correct_sample_id}'")

            processed_data.append(processed_row)
        
        # Store the uploaded data in the current experiment
        uploaded_data = {
            'filename': file.filename,
            'upload_date': datetime.now().isoformat(),
            'data': processed_data,
            'columns': df.columns.tolist(),
            'shape': [int(x) for x in df.shape],  # Convert tuple of numpy ints to list of Python ints
            'area_columns': area_columns  # Add area columns information
        }
        
        # Add to analytical data without overwriting existing data
        if 'analytical_data' not in current_experiment:
            current_experiment['analytical_data'] = {}
        
        # If analytical_data is a list (old format), convert it to new format
        if isinstance(current_experiment['analytical_data'], list):
            old_uploads = current_experiment['analytical_data']
            current_experiment['analytical_data'] = {
                'selectedCompounds': [],
                'uploadedFiles': old_uploads
            }
        
        if 'uploadedFiles' not in current_experiment['analytical_data']:
            current_experiment['analytical_data']['uploadedFiles'] = []
        
        current_experiment['analytical_data']['uploadedFiles'].append(uploaded_data)
        # Re-assign to trigger dirty flag (in-place mutation doesn't mark state dirty)
        current_experiment['analytical_data'] = current_experiment['analytical_data']

        logger.debug(f"Upload successful. Current experiment keys: {list(current_experiment.keys())}")
        
        return jsonify({
            'message': 'File uploaded successfully',
            'filename': file.filename,
            'rows': len(df),
            'columns': len(df.columns),
            'area_columns': area_columns
        }), 200
        
    except Exception as e:
        logger.error(f"Unexpected error in upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@uploads_bp.route('/materials/upload', methods=['POST'])
def upload_materials_from_excel():
    """Upload materials from Excel file"""
    try:
        logger.debug("Materials upload endpoint called")
        
        # Ensure inventory is loaded
        if not inventory_data:
            if not load_inventory():
                logger.warning("Warning: Could not load inventory data")
        
        if 'file' not in request.files:
            logger.debug("No file in request.files")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        logger.debug(f"File received: {file.filename}")
        
        if file.filename == '':
            logger.debug("Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'.xlsx', '.xls'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        logger.debug(f"File extension: {file_ext}")
        
        if file_ext not in allowed_extensions:
            return jsonify({'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'}), 400
        
        # Read the Excel file
        try:
            logger.debug("Attempting to read Excel file")
            excel_file = pd.ExcelFile(file)
            logger.debug(f"Excel sheets: {excel_file.sheet_names}")
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400
        
        # Look for Materials sheet
        if 'Materials' not in excel_file.sheet_names:
            return jsonify({'error': 'No "Materials" sheet found in the Excel file'}), 400
        
        # Read the Materials sheet
        try:
            materials_df = pd.read_excel(excel_file, sheet_name='Materials')
            logger.debug(f"Materials sheet read successfully. Shape: {materials_df.shape}")
        except Exception as e:
            logger.error(f"Error reading Materials sheet: {str(e)}")
            return jsonify({'error': f'Error reading Materials sheet: {str(e)}'}), 400
        
        # Extract materials from the sheet
        materials = []
        for index, row in materials_df.iterrows():
            # Skip empty rows
            if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                continue
            
            # Extract material data based on expected columns (support both old and new column names)
            # Helper function to clean and validate field values
            def clean_field(value):
                if pd.isna(value):
                    return ''
                str_value = str(value).strip()
                # Return empty string for common "empty" representations
                if str_value.lower() in ['nan', 'null', 'none', '']:
                    return ''
                return str_value
            
            material = {
                'name': clean_field(row.get('chemical_name', row.get('Chemical_Name', row.get('Name', row.iloc[1] if len(row) > 1 else '')))),
                'alias': clean_field(row.get('alias', row.get('Alias', ''))),
                'cas': clean_field(row.get('cas_number', row.get('CAS_Number', row.get('CAS', '')))),
                'smiles': clean_field(row.get('smiles', row.get('SMILES', ''))),
                'molecular_weight': clean_field(row.get('molecular_weight', row.get('Molecular_Weight', row.get('Molecular Weight', '')))),
                'barcode': clean_field(row.get('barcode', row.get('Barcode', row.get('Lot number', '')))),
                'role': clean_field(row.get('role', row.get('Role', ''))),
                'role_id': clean_field(row.get('role_id', row.get('Role_ID', ''))),
                'supplier': clean_field(row.get('supplier', row.get('Supplier', ''))),
                'catalog_number': clean_field(row.get('catalog_number', row.get('Catalog_Number', row.get('Cat #', '')))),
                'source': 'excel_upload'
            }
            
            # Only add if name is not empty
            if material.get('name') and material.get('name') != 'nan':
                materials.append(material)
        
        if not materials:
            return jsonify({'error': 'No valid materials found in the Materials sheet'}), 400
        
        logger.debug(f"Extracted {len(materials)} materials from Excel file")
        for i, mat in enumerate(materials):
            logger.debug(f"  Material {i+1}: {mat['name']} - CAS: {mat['cas']} - Alias: {mat['alias']}")
        
        # Get current materials
        current_materials = current_experiment.get('materials', [])
        
        # Check for duplicates and add new materials
        added_materials = []
        skipped_materials = []
        
        # First, check all materials against the original current_materials list
        for material in materials:
            logger.debug(f"Processing material: {material.get('name', 'Unknown')} (CAS: {material.get('cas', 'Unknown')})")
            # Check if material already exists in current experiment (by name, CAS, or SMILES)
            is_duplicate = any(
                (existing.get('name') and material.get('name') and existing.get('name') == material.get('name')) or
                (existing.get('cas') and material.get('cas') and existing.get('cas') == material.get('cas')) or
                (existing.get('smiles') and material.get('smiles') and existing.get('smiles') == material.get('smiles'))
                for existing in current_materials
            )
            
            if is_duplicate:
                logger.debug(f"  -> Skipping {material.get('name', 'Unknown')} (duplicate)")
                skipped_materials.append(material.get('alias') or material.get('name', 'Unknown'))
            else:
                # Check if material exists in inventory or private inventory by CAS number
                inventory_material = None
                if material.get('cas') and material.get('cas').strip():
                    # Check main inventory
                    if inventory_data:
                        inventory_match = inventory_data[
                            inventory_data['cas_number'].astype(str).str.strip() == material.get('cas').strip()
                        ]
                        if not inventory_match.empty:
                            inventory_material = inventory_match.iloc[0].to_dict()
                            logger.debug(f"  -> Found in main inventory: {inventory_material.get('chemical_name')}")
                        else:
                            logger.debug(f"  -> Not found in main inventory")
                    
                    # Check private inventory if not found in main inventory
                    if inventory_material is None:
                        config = get_config()
                        private_path = config.PRIVATE_INVENTORY_PATH
                        if os.path.exists(private_path):
                            try:
                                private_df = pd.read_excel(private_path, parse_dates=False)
                                # Convert all columns to string to avoid NaTType issues
                                for col in private_df.columns:
                                    private_df[col] = private_df[col].astype(str)
                                    private_df[col] = private_df[col].replace('nan', None)
                                
                                private_match = private_df[
                                    private_df['cas_number'].astype(str).str.strip() == material.get('cas').strip()
                                ]
                                if not private_match.empty:
                                    inventory_material = private_match.iloc[0].to_dict()
                                    logger.debug(f"  -> Found in private inventory: {inventory_material.get('chemical_name')}")
                                else:
                                    logger.debug(f"  -> Not found in private inventory")
                            except Exception as e:
                                logger.error(f"Error checking private inventory: {e}")
                
                # Use inventory data if found, otherwise use uploaded data
                if inventory_material:
                    logger.debug(f"  -> Using inventory data for {material.get('name', 'Unknown')}")
                    # Use inventory data but preserve some uploaded data
                    final_material = {
                        'name': inventory_material.get('chemical_name', material.get('name', '')),
                        'alias': material.get('alias', inventory_material.get('alias', '')),
                        'cas': inventory_material.get('cas_number', material.get('cas', '')),
                        'smiles': inventory_material.get('smiles', material.get('smiles', '')),
                        'molecular_weight': inventory_material.get('molecular_weight', material.get('molecular_weight', '')),
                        'barcode': material.get('barcode', inventory_material.get('barcode', '')),
                        'role': material.get('role', ''),
                        'role_id': material.get('role_id', ''),
                        'source': 'inventory_match',
                        'inventory_location': inventory_material.get('location', ''),
                        'supplier': material.get('supplier', inventory_material.get('supplier', '')),
                        'catalog_number': material.get('catalog_number', ''),
                    }
                else:
                    logger.debug(f"  -> Using uploaded data for {material.get('name', 'Unknown')}")
                    # Use uploaded data
                    final_material = material.copy()
                    final_material['source'] = 'excel_upload'
                
                added_materials.append(final_material)
        
        # Then, add all new materials to the current_materials list at once
        current_materials.extend(added_materials)
        
        # Update the experiment materials
        current_experiment['materials'] = current_materials
        
        # Count materials by source
        inventory_matches = len([m for m in added_materials if m.get('source') == 'inventory_match'])
        excel_uploads = len([m for m in added_materials if m.get('source') == 'excel_upload'])
        
        logger.debug(f"Final results: Added {len(added_materials)} materials ({inventory_matches} from inventory, {excel_uploads} from upload data)")
        for i, mat in enumerate(added_materials):
            logger.debug(f"  Added {i+1}: {mat.get('name', 'Unknown')} (source: {mat.get('source', 'Unknown')})")
        
        return jsonify({
            'message': 'Materials uploaded successfully',
            'filename': file.filename,
            'total_materials': len(materials),
            'added_materials': len(added_materials),
            'skipped_materials': len(skipped_materials),
            'inventory_matches': inventory_matches,
            'excel_uploads': excel_uploads,
            'added_material_names': [m.get('alias') or m.get('name', 'Unknown') for m in added_materials],
            'skipped_material_names': skipped_materials
        }), 200
        
    except Exception as e:
        logger.error(f"Unexpected error in materials upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Materials upload failed: {str(e)}'}), 500

@uploads_bp.route('/procedure/update-plate-type', methods=['POST'])
def update_procedure_plate_type():
    """Update the procedure plate type when kit is applied"""
    try:
        data = request.json
        new_plate_type = data.get('plate_type', '96')
        current_procedure = data.get('current_procedure', [])
        
        logger.debug(f"Updating procedure plate type to: {new_plate_type}")
        
        # Store the plate type information in the experiment context
        if 'context' not in current_experiment:
            current_experiment['context'] = {}
        
        current_experiment['context']['plate_type'] = new_plate_type
        
        # Validate that all wells in current procedure fit in the new plate type
        plate_configs = {
            '24': {'max_row': 'D', 'max_col': 6},
            '48': {'max_row': 'F', 'max_col': 8},
            '96': {'max_row': 'H', 'max_col': 12}
        }
        
        plate_config = plate_configs.get(new_plate_type, plate_configs['96'])
        
        # Filter out any procedure entries that don't fit in the new plate
        valid_procedure = []
        for entry in current_procedure:
            well = entry.get('well', '')
            if well and len(well) >= 2:
                row_letter = well[0]
                col_number = int(well[1:]) if well[1:].isdigit() else 0
                
                if (row_letter <= plate_config['max_row'] and 
                    col_number <= plate_config['max_col']):
                    valid_procedure.append(entry)
        
        # Update the procedure with valid entries
        current_experiment['procedure'] = valid_procedure
        
        return jsonify({
            'success': True, 
            'plate_type': new_plate_type,
            'filtered_wells': len(current_procedure) - len(valid_procedure)
        })
        
    except Exception as e:
        logger.error(f"Error updating procedure plate type: {str(e)}")
        return jsonify({'error': f'Failed to update plate type: {str(e)}'}), 500
