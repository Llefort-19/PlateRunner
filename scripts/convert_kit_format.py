"""
Kit file format converter: Wide → Long format.

Converts existing kit .xlsx files from old wide-format Design sheet
to the new long-format Design sheet.

Usage:
    python convert_kit_format.py input.xlsx                 # saves to input_converted.xlsx
    python convert_kit_format.py input.xlsx output.xlsx     # saves to output.xlsx
    python convert_kit_format.py input.xlsx --in-place      # overwrites input file
"""
import sys
import os
import argparse
from pathlib import Path

from openpyxl import load_workbook, Workbook


def convert_kit(input_path: str, output_path: str) -> None:
    """Convert a kit Excel file from wide to long Design format."""
    wb = load_workbook(input_path)

    if 'Design' not in wb.sheetnames:
        print(f"ERROR: No 'Design' sheet in {input_path}")
        sys.exit(1)

    # Read Materials for Nr lookup
    materials = []
    if 'Materials' in wb.sheetnames:
        ws_mat = wb['Materials']
        mat_headers = [str(c.value or '').strip().lower() for c in ws_mat[1]]
        for row in ws_mat.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            mat = {}
            for i, h in enumerate(mat_headers):
                val = row[i] if i < len(row) else None
                if val is not None:
                    mat[h] = str(val).strip()
            materials.append(mat)
    print(f"Loaded {len(materials)} materials")

    # Build name/alias → Nr lookup
    name_to_nr = {}
    alias_to_nr = {}
    for i, m in enumerate(materials):
        nr = i + 1
        name = (m.get('chemical_name') or m.get('name', '')).strip().lower()
        alias = (m.get('alias', '')).strip().lower()
        if name:
            name_to_nr[name] = nr
        if alias:
            alias_to_nr[alias] = nr

    # Read old Design sheet
    ws_design = wb['Design']
    headers = [str(c.value or '').strip() for c in ws_design[1]]

    # Determine well column
    well_col = None
    for i, h in enumerate(headers):
        if h.lower() in ('well', 'well_id'):
            well_col = i
            break
    if well_col is None:
        well_col = 0  # Assume first column

    # Find paired columns (name, amount)
    # Old format has: Well, ID, Compound 1 Name, Compound 1 Amount, ...
    paired_cols = []
    i = 2  # Skip Well and ID
    while i + 1 < len(headers):
        name_h = headers[i]
        amt_h = headers[i + 1]
        if ('name' in name_h.lower() or 'Name' in name_h) and \
           ('amount' in amt_h.lower() or 'Amount' in amt_h):
            paired_cols.append((i, i + 1, name_h))
        i += 2

    if not paired_cols:
        print("WARNING: Could not detect paired (name, amount) columns. Trying positional.")
        i = 2
        while i + 1 < len(headers):
            paired_cols.append((i, i + 1, headers[i]))
            i += 2

    print(f"Found {len(paired_cols)} compound columns")

    # Build long-format rows
    long_rows = []

    for row in ws_design.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        well = str(row[well_col] or '').strip()
        if not well or well == 'nan':
            continue

        dispense_order = 0
        for name_idx, amt_idx, col_header in paired_cols:
            name_val = str(row[name_idx] or '').strip() if name_idx < len(row) else ''
            amt_val = str(row[amt_idx] or '').strip() if amt_idx < len(row) else ''

            if name_val and name_val != 'nan' and amt_val and amt_val != 'nan':
                dispense_order += 1
                # Resolve Nr
                name_lower = name_val.lower()
                nr = alias_to_nr.get(name_lower) or name_to_nr.get(name_lower) or ''

                long_rows.append([well, nr, name_val, amt_val, dispense_order])

    print(f"Converted {len(long_rows)} material entries across wells")

    # Remove old Design sheet and create new one
    del wb['Design']
    ws_new = wb.create_sheet('Design', index=wb.sheetnames.index('Materials') + 1 if 'Materials' in wb.sheetnames else 1)

    ws_new.append(['well', 'material_nr', 'material_alias', 'amount', 'dispense_order'])
    for r in long_rows:
        ws_new.append(r)

    # Save
    wb.save(output_path)
    print(f"Saved converted file to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Convert kit .xlsx from wide to long Design format')
    parser.add_argument('input', help='Input .xlsx file')
    parser.add_argument('output', nargs='?', default=None, help='Output .xlsx file (default: input_converted.xlsx)')
    parser.add_argument('--in-place', action='store_true', help='Overwrite the input file')

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: File not found: {args.input}")
        sys.exit(1)

    if args.in_place:
        output = args.input
    elif args.output:
        output = args.output
    else:
        stem = Path(args.input).stem
        ext = Path(args.input).suffix
        output = str(Path(args.input).parent / f"{stem}_converted{ext}")

    convert_kit(args.input, output)


if __name__ == '__main__':
    main()
